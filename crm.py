import sys
from PyQt5.QtWidgets import QApplication,QHeaderView,QSpinBox,QAbstractItemView,QTimeEdit,QDateEdit,QShortcut,QDesktopWidget,QDialog,QMainWindow,QDateTimeEdit, QLabel, QComboBox,QGridLayout,QDialog, QWidget, QTableWidget, QTableWidgetItem, QPushButton, QVBoxLayout, QHBoxLayout, QLineEdit, QMessageBox, QTabWidget,QFileDialog,QProgressDialog
from PyQt5.QtCore import Qt,QDate,QTimer,QUrl
from PyQt5.QtGui import QFont,QKeySequence,QIcon,QColor, QBrush,QDesktopServices
import pymysql
import hashlib
import shutil
import re
import openpyxl
from datetime import datetime
import os,time
import requests
import configparser

#版本号
CURRENT_VERSION = "1.0"
#检测升级URL
SERVER_URL = "" #你的升级链接
#会员等级
LEVELS = ["普通卡", "银卡", "金卡", "钻卡", "至尊卡"]

search_entry = None
search_table = None
transaction_table = None
transaction_member_entry = None
#打包命令
#pyinstaller --onefile --noconsole --icon=icon.ico your_code.py

#检查更新
def check_for_updates():
    try:
        response = requests.get(f"{SERVER_URL}/version.txt")
        latest_version = response.text.strip()
        if latest_version > CURRENT_VERSION:
            return latest_version
        return None
    except Exception as e:
        print(f"检查更新时出错: {e}")
        return None

#下载更新文件
def download_new_version(latest_version, progress_callback=None):
    new_version_url = f"{SERVER_URL}/crm_v{latest_version}.exe"
    filename = "CRM_V%s.exe" % latest_version
    try:
        response = requests.get(new_version_url, stream=True)
        total_length = response.headers.get('content-length')
        
        with open(filename, 'wb') as f:
            if total_length is None:  # 没有内容长度
                f.write(response.content)
            else:
                downloaded = 0
                total_length = int(total_length)
                for chunk in response.iter_content(chunk_size=8192):
                    if progress_callback and not progress_callback(downloaded / total_length):
                        return "canceled"

                    downloaded += len(chunk)
                    f.write(chunk)
        return True
    except Exception as e:
        print(f"下载新版本时出错: {e}")
        return False
# 登录界面
class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        # 设置窗口图标
        if getattr(sys, 'frozen', False):
            # 如果是打包状态
            bundle_dir = sys._MEIPASS
        else:
            # 如果是正常运行状态
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        ICON_PATH = os.path.join(bundle_dir, 'icon.ico')
        self.db_manager = DatabaseManager()
        self.setWindowIcon(QIcon(ICON_PATH))
        # 窗口属性
        self.setWindowTitle("会员管理系统")
        self.setGeometry(300, 300, 350, 200)
        self.center_on_screen()
        layout = QVBoxLayout()

        # 标题标签
        title_label = QLabel("登录")
        title_label.setStyleSheet("font-size: 32px;padding:10px;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # 用户名输入
        self.username_entry = QLineEdit()
        self.username_entry.setPlaceholderText("用户名")
        layout.addWidget(self.username_entry)
        
        # 密码输入
        self.password_entry = QLineEdit()
        self.password_entry.setPlaceholderText("密码")
        self.password_entry.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_entry)
        
        # 登录按钮
        login_button = QPushButton("登录")
        login_button.clicked.connect(self.login)
        self.password_entry.returnPressed.connect(self.login)  # 这行代码将回车键与登录按钮连接
        layout.addWidget(login_button)
        
        self.setLayout(layout)
    def center_on_screen(self):
        screen = QDesktopWidget().screenGeometry()  # 获取屏幕的尺寸
        window = self.geometry()  # 获取窗口的尺寸
        self.move((screen.width() - window.width()) / 2,(screen.height() - window.height()) / 2)

    def login(self):
        username = self.username_entry.text()
        password = self.password_entry.text()

        # 查询用户
        query = "SELECT id, password_hash, salt FROM users WHERE username = %s"
        user = self.db_manager.execute_query(query, (username,))

        if user is None:
            # 用户名不存在
            QMessageBox.warning(self, "警告", "用户名或密码错误！")
        else:
            user_id, password_hash, salt = user[0]

            # 生成密码的哈希值
            salted_password = password + salt
            hash_object = hashlib.sha256(salted_password.encode())
            hashed_password = hash_object.hexdigest()
            if hashed_password == password_hash:
                # 密码正确，登录成功
                self.close()  # 关闭登录窗口
                self.main_window = MainWindow(user_id)  # 创建主窗口，传入用户ID
                self.main_window.show()  # 显示主窗口
            else:
                # 密码错误
                QMessageBox.warning(self, "警告", "用户名或密码错误！")


#主界面
class MainWindow(QMainWindow):
    def __init__(self,user_id):
        super().__init__()
        if getattr(sys, 'frozen', False):
            # 如果是打包状态
            bundle_dir = sys._MEIPASS
        else:
            # 如果是正常运行状态
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        ICON_PATH = os.path.join(bundle_dir, 'icon.ico')
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setWindowTitle("会员管理系统V1.0")
        self.setGeometry(30, 30, 1300, 720)
        self.center_on_screen()

        self.db_manager = DatabaseManager()
        self.creator_id = user_id

        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        # 创建Tab控件
        self.tab_widget = QTabWidget()

        # 创建会员查询界面
        search_widget = QWidget()
        search_layout = QVBoxLayout()
        
        # 会员查询界面
        search_label = QLabel("请输入会员姓名或手机号：")
        self.search_entry = QLineEdit()
        self.search_entry.returnPressed.connect(self.search_member)
        search_button = QPushButton("查询[F5]")

        search_button.clicked.connect(self.search_member)
        

        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_entry)
        search_layout.addWidget(search_button)

        # 创建按钮布局
        button_layout = QHBoxLayout()

        add_button = QPushButton("新增会员")
        add_button.clicked.connect(self.create_add_member_window)
        edit_button = QPushButton("修改会员信息")
        edit_button.clicked.connect(self.edit_member)
        delete_button = QPushButton("删除会员")
        delete_button.clicked.connect(self.delete_member)
        recharge_button = QPushButton("充值")
        recharge_button.clicked.connect(self.recharge_dialog)
        buy_product_button = QPushButton("购买商品")
        buy_product_button.clicked.connect(self.buy_product)
        transaction_button = QPushButton("会员消费")
        transaction_button.clicked.connect(self.create_consume_dialog)
        buy_beauty_project_button = QPushButton("购买美容项目")
        buy_beauty_project_button.clicked.connect(self.buy_beauty_project)
        #设置CSS样式
        font = QFont()
        font.setBold(True)

        search_label.setStyleSheet("color: red;")
        search_label.setFont(font)
        search_button.setFont(font)
        add_button.setFont(font)
        edit_button.setFont(font)
        delete_button.setFont(font)
        recharge_button.setFont(font)
        transaction_button.setFont(font)
        buy_beauty_project_button.setFont(font)

        button_layout.addWidget(add_button)
        button_layout.addWidget(edit_button)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(recharge_button)
        button_layout.addWidget(buy_product_button)
        button_layout.addWidget(transaction_button)
        button_layout.addWidget(buy_beauty_project_button)
        
        self.search_table = QTableWidget()

        self.search_table.setColumnCount(9)
        self.search_table.setHorizontalHeaderLabels([ "姓名", "手机号", "生日", "注册日期", "最后消费日期", "余额", "会员等级","ID","累计充值"])
        self.search_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.search_table.doubleClicked.connect(self.edit_member) #双击事件
                        
        # 设置所有列为Interactive模式，以允许手动调整列宽
        self.search_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        self.search_table.setColumnHidden(7, True)  # 隐藏第八列（索引从0开始计数）

        search_layout.addLayout(button_layout)
        search_layout.addWidget(self.search_table)

        search_widget.setLayout(search_layout)

        # 创建充值/消费记录界面
        transaction_widget = QWidget()

        transaction_search_layout = QVBoxLayout()
        transaction_label = QLabel("请输入会员姓名/手机号/营业员姓名：")
        transaction_label.setStyleSheet("color: red;")
        transaction_label.setFont(font)
        self.transaction_member_entry = QLineEdit()
        self.transaction_member_entry.returnPressed.connect(self.search_transaction)
        transaction_button = QPushButton("查询[F5]")
        transaction_button.clicked.connect(self.search_transaction)
        
        transaction_search_layout.addWidget(transaction_label)
        transaction_search_layout.addWidget(self.transaction_member_entry)
        transaction_search_layout.addWidget(transaction_button)
        
        # 创建按钮布局
        transaction_button_layout = QHBoxLayout()
        export_button = QPushButton("导出")
        export_button.clicked.connect(self.export_transaction_table)
        reverse_button = QPushButton("冲账")
        reverse_button.clicked.connect(self.reverse_transaction)
        transaction_button_layout.addWidget(reverse_button)
        transaction_button_layout.addWidget(export_button)

        self.transaction_table = QTableWidget()
        self.transaction_table.setColumnCount(9)
        self.transaction_table.setHorizontalHeaderLabels(["姓名", "手机号", "金额", "时间", "项目", "备注", "余额","营业员","会员等级"])

        # 在这里设置选择行为
        self.transaction_table.setSelectionBehavior(QTableWidget.SelectRows)

        self.transaction_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        transaction_search_layout.addLayout(transaction_button_layout)
        transaction_search_layout.addWidget(self.transaction_table)

        transaction_widget.setLayout(transaction_search_layout)
        # 创建美容项目管理界面
        beauty_project_widget = QWidget()
        beauty_project_layout = QVBoxLayout()
        # 美容项目管理界面
        beauty_project_label = QLabel("请输入美容项目名称：")
        beauty_project_label.setStyleSheet("color: red;")
        beauty_project_label.setFont(font)
        self.beauty_project_entry = QLineEdit()
        self.beauty_project_entry.returnPressed.connect(self.search_beauty_project)
        beauty_project_button = QPushButton("查询[F5]")
        beauty_project_button.clicked.connect(self.search_beauty_project)

        beauty_project_layout.addWidget(beauty_project_label)
        beauty_project_layout.addWidget(self.beauty_project_entry)
        beauty_project_layout.addWidget(beauty_project_button)
        # 创建按钮布局
        beauty_project_button_layout = QHBoxLayout()

        add_beauty_project_button = QPushButton("新增美容项目")
        add_beauty_project_button.clicked.connect(self.create_add_beauty_project_window)
        edit_beauty_project_button = QPushButton("修改美容项目信息")
        edit_beauty_project_button.clicked.connect(self.edit_beauty_project)
        delete_beauty_project_button = QPushButton("删除美容项目")
        delete_beauty_project_button.clicked.connect(self.delete_beauty_project)

        beauty_project_button_layout.addWidget(add_beauty_project_button)
        beauty_project_button_layout.addWidget(edit_beauty_project_button)
        beauty_project_button_layout.addWidget(delete_beauty_project_button)

        self.beauty_project_table = QTableWidget()

        self.beauty_project_table.setColumnCount(6)
        self.beauty_project_table.setHorizontalHeaderLabels(["ID", "美容项目名称", "套餐次数", "单次价格", "套餐价格", "单次耗时", "新建时间"])
        self.beauty_project_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.beauty_project_table.doubleClicked.connect(self.edit_beauty_project) #双击事件

        self.beauty_project_table.setColumnHidden(0, True)  # 隐藏第1列（索引从0开始计数）
        self.beauty_project_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        beauty_project_layout.addLayout(beauty_project_button_layout)
        beauty_project_layout.addWidget(self.beauty_project_table)

        beauty_project_widget.setLayout(beauty_project_layout)

        # 创建营业员管理界面
        clerks_widget = QWidget()
        clerks_layout = QVBoxLayout()
        # 营业员管理界面
        clerks_label = QLabel("请输入营业员名称：")
        clerks_label.setStyleSheet("color: red;")
        clerks_label.setFont(font)
        self.clerks_entry = QLineEdit()
        self.clerks_entry.returnPressed.connect(self.search_clerk)
        clerks_button = QPushButton("查询[F5]")
        clerks_button.clicked.connect(self.search_clerk)

        clerks_layout.addWidget(clerks_label)
        clerks_layout.addWidget(self.clerks_entry)
        clerks_layout.addWidget(clerks_button)
        # 创建按钮布局
        clerks_button_layout = QHBoxLayout()

        add_clerks_button = QPushButton("新增营业员")
        add_clerks_button.clicked.connect(self.create_add_clerk_window)
        change_clerks_button = QPushButton("修改营业员")
        change_clerks_button.clicked.connect(self.change_clerk_window)

        delete_clerks_button = QPushButton("删除营业员")
        delete_clerks_button.clicked.connect(self.delete_clerks)

        clerks_button_layout.addWidget(add_clerks_button)
        clerks_button_layout.addWidget(change_clerks_button)
        clerks_button_layout.addWidget(delete_clerks_button)

        self.clerks_table = QTableWidget()

        self.clerks_table.setColumnCount(3)
        self.clerks_table.setHorizontalHeaderLabels(["ID", "营业员名称","手机号"])
        self.clerks_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.clerks_table.doubleClicked.connect(self.change_clerk_window) #双击事件

        self.clerks_table.setColumnHidden(0, True)  # 隐藏第1列（索引从0开始计数）
        self.clerks_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        clerks_layout.addLayout(clerks_button_layout)
        clerks_layout.addWidget(self.clerks_table)

        clerks_widget.setLayout(clerks_layout)

        # 创建库存管理界面
        inventory_management_widget = QWidget()
        inventory_layout = QVBoxLayout()

        # 库存管理界面标签和输入框
        inventory_label = QLabel("请输入商品条码或名称：")
        inventory_label.setStyleSheet("color: blue;")
        self.inventory_entry = QLineEdit()
        # 假设有一个名为search_inventory的方法来进行搜索操作
        self.inventory_entry.returnPressed.connect(self.search_inventory)
        inventory_button = QPushButton("查询库存")
        inventory_button.clicked.connect(self.search_inventory)

        inventory_layout.addWidget(inventory_label)
        inventory_layout.addWidget(self.inventory_entry)
        inventory_layout.addWidget(inventory_button)

        # 创建按钮布局
        inventory_button_layout = QHBoxLayout()

        import_inventory_button = QPushButton("导入库存")
        # 导入库存操作
        import_inventory_button.clicked.connect(self.open_import_inventory_dialog)

        inventory_audit_button = QPushButton("盘点库存")
        # 盘点库存操作
        inventory_audit_button.clicked.connect(self.open_inventoryCheck_dialog)

        query_product_flow_button = QPushButton("查询单品流水")
        # 查询单品流水操作
        query_product_flow_button.clicked.connect(self.query_product_flow)

        inventory_button_layout.addWidget(import_inventory_button)
        inventory_button_layout.addWidget(inventory_audit_button)
        inventory_button_layout.addWidget(query_product_flow_button)


        self.inventory_table = QTableWidget()

        self.inventory_table.setColumnCount(5)
        self.inventory_table.setHorizontalHeaderLabels(["ID", "商品条码", "商品名称", "库存数量","最后更新时间"])
        self.inventory_table.setSelectionBehavior(QTableWidget.SelectRows)

        self.inventory_table.setColumnHidden(0, True)  # 隐藏ID列
        self.inventory_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        inventory_layout.addLayout(inventory_button_layout)
        inventory_layout.addWidget(self.inventory_table)

        inventory_management_widget.setLayout(inventory_layout)

        # 假设self.tab_widget是您的QTabWidget实例
        self.tab_widget.addTab(inventory_management_widget, "库存管理")

        # 创建系统设置界面
        password_widget = QWidget()
        password_layout = QGridLayout()

        # 定义会员等级折扣设置按钮
        set_discount_button = QPushButton("会员等级折扣设置")
        set_discount_button.clicked.connect(self.show_discount_settings_dialog)

        # 定义会员等级规则设置按钮
        set_level_rules_button = QPushButton("会员等级规则设置")
        set_level_rules_button.clicked.connect(self.show_level_rules_dialog)  # 这里需要定义一个新的槽函数来处理这个事件

        # 定义会员日设置按钮
        set_birthday_rules_button = QPushButton("会员日设置")
        set_birthday_rules_button.clicked.connect(self.show_birthday_rules_dialog)  # 这里定义一个新的槽函数来处理这个事件

        # 定义营业时间设置按钮
        set_business_hours_button = QPushButton("营业时间设置")
        set_business_hours_button.clicked.connect(self.show_business_hours_dialog)  # 这里定义一个新的槽函数来处理这个事件

        # 定义美容床位设置按钮
        set_beauty_bed_button = QPushButton("美容床位设置")
        set_beauty_bed_button.clicked.connect(self.show_beauty_bed_dialog)  # 这里需要定义一个新的槽函数来处理这个事件

        change_password_button = QPushButton("修改密码")
        change_password_button.clicked.connect(self.show_change_password_dialog)

        logout_button = QPushButton("注销")
        logout_button.clicked.connect(self.logout)

        exit_button = QPushButton("退出")
        exit_button.clicked.connect(QApplication.instance().quit)

        password_layout.addWidget(set_discount_button, 0, 0)
        password_layout.addWidget(set_level_rules_button, 0, 1)
        password_layout.addWidget(set_birthday_rules_button, 1, 0)
        password_layout.addWidget(set_business_hours_button, 1, 1)
        password_layout.addWidget(set_beauty_bed_button, 2, 0)  # 新增的美容床位设置按钮
        password_layout.addWidget(change_password_button, 2, 1)
        password_layout.addWidget(logout_button, 3, 0)
        password_layout.addWidget(exit_button, 3, 1)


        password_widget.setLayout(password_layout)

        self.tab_widget.addTab(password_widget, "修改密码")


        # 创建会员美容项目次数报表界面
        report_widget = QWidget()
        report_layout = QVBoxLayout()
        report_label = QLabel("请输入会员姓名或手机号：")
        report_label.setStyleSheet("color: red;")
        report_label.setFont(font)
        self.report_entry = QLineEdit()
        self.report_entry.returnPressed.connect(self.search_report)
        report_button = QPushButton("查询[F5]")
        report_button.clicked.connect(self.search_report)

        self.report_table = QTableWidget()
        self.report_table.setColumnCount(4)
        self.report_table.setHorizontalHeaderLabels(["会员姓名", "手机号码","美容项目", "剩余次数"])
        self.report_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        report_layout.addWidget(report_label)
        report_layout.addWidget(self.report_entry)
        report_layout.addWidget(report_button)
        report_layout.addWidget(self.report_table)

        report_widget.setLayout(report_layout)
        #会员预约查询选项卡
        appointment_viewer_tab = AppointmentViewerTab(self.db_manager,self.creator_id)

        self.tab_widget.addTab(search_widget, "会员查询")
        self.tab_widget.addTab(appointment_viewer_tab, "预约查看")
        self.tab_widget.addTab(transaction_widget, "会员充值/消费记录")
        self.tab_widget.addTab(report_widget, "会员项目剩余次数查询")
        self.tab_widget.addTab(beauty_project_widget, "美容项目设置")
        self.tab_widget.addTab(clerks_widget, "营业员设置")
        self.tab_widget.addTab(inventory_management_widget, "库存管理")
        self.tab_widget.addTab(password_widget, "系统设置")
        # 创建全局快捷键
        search_shortcut = QShortcut(QKeySequence("F5"), self)
        search_shortcut.activated.connect(self.global_search) # 连接到全局查询方法

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tab_widget)
        
        main_widget.setLayout(main_layout)
        self.sort_order = Qt.AscendingOrder  # 初始排序顺序为升序
        # 点击充值/消费记录表头进行排序
        self.transaction_table.horizontalHeader().sectionClicked.connect(self.sort_transaction_table)
    #打开单品流水对话框
    def query_product_flow(self):
        selected_row = self.inventory_table.currentRow()
        if selected_row != -1:
            barcode = self.inventory_table.item(selected_row, 1).text()
            Dialog = QueryProductFlowDialog(barcode,self.creator_id,self)
            Dialog.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要查询的商品。")
    def adjustColumnWidthsToContents(self):
        for column in range(self.search_table.columnCount()):
            self.search_table.resizeColumnToContents(column)
    def center_on_screen(self):
        screen = QDesktopWidget().screenGeometry()  # 获取屏幕的尺寸
        window = self.geometry()  # 获取窗口的尺寸
        self.move((screen.width() - window.width()) / 2,(screen.height() - window.height()) / 2)
    #实现F5在当前选项卡执行查询
    def global_search(self):
        # 检查当前选中的选项卡，并调用相应的查询方法
        current_tab = self.tab_widget.currentIndex() # 获取当前选中的选项卡索引
        if current_tab == 0:
            self.search_member()
        elif current_tab == 2:
            self.search_transaction()
        elif current_tab == 3:
            self.search_report()
        elif current_tab == 4:
            self.search_beauty_project()
        elif current_tab == 5:
            self.search_clerk()

    # 会员查询功能
    def search_member(self):
        keyword = self.search_entry.text()
        query = """
            SELECT members.name, members.phone, members.birthday, members.register_date, MAX(transactions.transaction_date), members.balance, members.membership_level, members.id, members.total_recharge_amount
            FROM members
            LEFT JOIN transactions ON transactions.member_id = members.id
            WHERE (members.name LIKE %s OR members.phone LIKE %s) AND members.creator_id = %s
            GROUP BY members.id
            ORDER BY MAX(transactions.transaction_date) DESC
        """
        params = ('%' + keyword + '%', '%' + keyword + '%', self.creator_id)
        members = self.db_manager.execute_query(query, params)

        self.search_table.setRowCount(0)
        for row, member in enumerate(members):
            self.search_table.insertRow(row)
            for col, data in enumerate(member):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled) #高亮修改这里
                self.search_table.setItem(row, col, item)
        # 如果有结果，选择第一行
        self.adjustColumnWidthsToContents()
        if self.search_table.rowCount() > 0:
            self.search_table.setCurrentCell(0, 0)

    # 会员新增
    def create_add_member_window(self):
        add_member_window = AddMemberWindow(self.creator_id, self)
        add_member_window.exec_()
    # 会员修改
    def edit_member(self):
        selected_row = self.search_table.currentRow()
        if selected_row != -1:
            member_id = self.search_table.item(selected_row, 7).text()
            edit_member_window = EditMemberWindow(member_id, self)
            edit_member_window.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要修改的会员。")
    # 会员删除
    def delete_member(self):
        selected_row = self.search_table.currentRow()
        if selected_row != -1:
            member_id = self.search_table.item(selected_row, 7).text()
            query = f"SELECT * FROM members WHERE id = {member_id}"
            member = self.db_manager.execute_query(query)[0]
            if member:
                # 检查是否存在对应的交易记录
                query = f"SELECT * FROM transactions WHERE member_id = {member_id}"
                transactions = self.db_manager.execute_query(query)
                if transactions:
                    QMessageBox.warning(self, "错误", "该会员已有消费记录，无法删除。")
                else:
                    confirm = QMessageBox.question(self, "确认", "【警告】确定要删除该会员吗？")
                    if confirm == QMessageBox.Yes:
                        query = f"DELETE FROM members WHERE id = {member_id}"
                        self.db_manager.execute_query(query)
                        QMessageBox.information(self, "提示", "会员删除成功。")
                        self.search_member()
            else:
                QMessageBox.warning(self, "错误", "选择的会员不存在。")
        else:
            QMessageBox.warning(self, "警告", "请选择要删除的会员。")
    # 会员充值
    def recharge_dialog(self):
        selected_row = self.search_table.currentRow()
        if selected_row != -1:
            member_id = self.search_table.item(selected_row, 7).text()
            recharge_dialog = RechargeDialog(member_id, self)
            recharge_dialog.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要充值的会员。")
    # 会员购买商品
    def buy_product(self):
        selected_row = self.search_table.currentRow()
        if selected_row != -1:
            member_id = self.search_table.item(selected_row, 7).text()
            consume_dialog = PurchaseProductDialog(member_id, parent=self)
            consume_dialog.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要消费的会员。")

    # 会员消费
    def create_consume_dialog(self):
        selected_row = self.search_table.currentRow()
        if selected_row != -1:
            member_id = self.search_table.item(selected_row, 7).text()
            consume_dialog = ConsumeDialog(member_id, parent=self, appointment_id=None, beauty_project_id=None)
            consume_dialog.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要消费的会员。")
    # 会员购卡
    def buy_beauty_project(self):
        selected_row = self.search_table.currentRow()
        if selected_row != -1:
            member_id = self.search_table.item(selected_row, 7).text()
            buy_beauty_project_dialog = BuyBeautyProjectDialog(member_id, self)
            buy_beauty_project_dialog.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要购买美容项目的会员。")

    # 会员充值消费记录查询
    def search_transaction(self):
        keyword = self.transaction_member_entry.text()
        query = """
        SELECT transactions.id, members.name, members.phone, transactions.amount, DATE_FORMAT(transactions.transaction_date, '%%Y-%%m-%%d %%H:%%i:%%s'), transactions.project, transactions.remark, transactions.balance, clerks.name, transactions.membership_level
        FROM transactions
        LEFT JOIN members ON transactions.member_id = members.id
        LEFT JOIN clerks ON transactions.clerk_id = clerks.id
        WHERE (members.name LIKE %s OR members.phone LIKE %s OR clerks.name LIKE %s) AND members.creator_id = %s
        ORDER BY transactions.transaction_date DESC
        """

        params = (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', self.creator_id)
        transactions = self.db_manager.execute_query(query, params)
        self.transaction_table.setRowCount(0)
        for row, transaction in enumerate(transactions):
            self.transaction_table.insertRow(row)
            # 从查询结果中获取交易ID，并保存为自定义数据
            transaction_id = transaction[0]
            for col, data in enumerate(transaction[1:]): # 注意这里从1开始，跳过交易ID
                item = QTableWidgetItem(str(data))
                item.setData(Qt.UserRole, transaction_id) # 保存交易ID作为自定义数据
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled) # 高亮修改这里
                self.transaction_table.setItem(row, col, item)

    #导出会员充值/消费记录
    def export_transaction_table(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 添加表头
        headers = ["姓名", "手机号", "金额", "时间", "项目", "备注", "余额", "营业员","会员等级"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # 添加数据
        keyword = self.transaction_member_entry.text()
        query = """
        SELECT members.name, members.phone, transactions.amount, transactions.transaction_date, transactions.project, transactions.remark, transactions.balance, clerks.name, transactions.membership_level
        FROM transactions
        LEFT JOIN members ON transactions.member_id = members.id
        LEFT JOIN clerks ON transactions.clerk_id = clerks.id
        WHERE (members.name LIKE %s OR members.phone LIKE %s OR clerks.name LIKE %s) AND members.creator_id = %s
        ORDER BY transactions.transaction_date DESC
        """

        params = (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', self.creator_id)
        transactions = self.db_manager.execute_query(query, params)
        for row_num, transaction in enumerate(transactions, 2):
            for col_num, value in enumerate(transaction, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # 选择保存路径
        filename, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel Files (*.xlsx)")
        if filename:
            workbook.save(filename)
            QMessageBox.information(self, "提示", f"会员充值/消费记录表已成功导出为：{filename}")
    #冲账
    def reverse_transaction(self):
        selected_row = self.transaction_table.currentRow()
        if selected_row != -1:
            confirm = QMessageBox.question(self, "确认", "您确定要冲账此交易吗？")
            if confirm == QMessageBox.Yes:
                self.perform_reverse_transaction(selected_row)
        else:
            QMessageBox.warning(self, "警告", "请选择要冲账的交易。")
    def get_member_id_by_transaction_id(self, transaction_id):
        query = "SELECT member_id FROM transactions WHERE id = %s"
        result = self.db_manager.execute_query(query, (transaction_id,))
        if result:
            return result[0][0]
        else:
            return None
    def update_member_balance(self, member_id, amount_change):
        # 更新会员余额
        query = "UPDATE members SET balance = balance + %s WHERE id = %s"
        self.db_manager.execute_query(query, (amount_change, member_id))
        
    def perform_reverse_transaction(self, row):
        # 获取交易ID
        item = self.transaction_table.item(row, 0)
        transaction_id = int(item.data(Qt.UserRole))
        # 开始事务
        self.db_manager.begin_transaction()

        try:
            # 获取所需冲账的交易记录
            query = "SELECT member_id, amount, project, clerk_id, membership_level, phone,remark FROM transactions WHERE id = %s"
            transaction_details = self.db_manager.execute_query(query, (transaction_id,))
            current_text = transaction_details[0][2]
            product_info = current_text.split("|")
            barcode =  product_info[0]
            product_name = product_info[1].split("*")[0]
            change_quantity = int(product_info[1].split("*")[1])
            query = "SELECT inventory_quantity FROM product_inventory WHERE barcode = %s AND creator_id = %s"
            old_inventory = self.db_manager.execute_query(query, (barcode,self.creator_id))

            if not transaction_details:
                raise Exception("交易未找到")

            member_id, original_amount, project, clerk_id, membership_level, phone,old_remark = transaction_details[0]
            
            if old_inventory:
                if old_remark == "冲账":
                    old_inventory = old_inventory[0][0]
                    new_inventory =  old_inventory - change_quantity
                else:
                    old_inventory = old_inventory[0][0]
                    new_inventory =  old_inventory + change_quantity
            if "已冲账" in old_remark:
                raise Exception("错误！该交易已被冲账,请检查")
            else:
                # 获取会员当前余额
                query = "SELECT balance FROM members WHERE id = %s"
                member_balance_result = self.db_manager.execute_query(query, (member_id,))
                if not member_balance_result:
                    raise Exception("会员未找到")

                member_balance = member_balance_result[0][0]
                reverse_balance = member_balance - original_amount
                membership_level = membership_level
                if old_remark == '购买现付':
                    # 创建冲账记录
                    reverse_query = """
                    INSERT INTO transactions (member_id, amount, transaction_date, remark, project, clerk_id, reversed, membership_level, phone)
                    VALUES (%s, %s, NOW(), '冲账', %s, %s, FALSE, %s, %s)
                    """
                    inserted_id  = self.db_manager.execute_insert(reverse_query,(member_id, -original_amount, project, clerk_id ,membership_level, phone))
                else:
                    # 创建冲账记录
                    reverse_query = """
                    INSERT INTO transactions (member_id, amount, transaction_date, remark, project, balance, clerk_id, reversed, membership_level, phone)
                    VALUES (%s, %s, NOW(), '冲账', %s, %s, %s, FALSE, %s, %s)
                    """
                    inserted_id  = self.db_manager.execute_insert(reverse_query, (member_id, -original_amount, project, reverse_balance, clerk_id ,membership_level, phone))
                    # 更新会员余额
                    self.update_member_balance(member_id, -original_amount)
                if old_inventory:
                    self.db_manager.execute_query("UPDATE product_inventory SET inventory_quantity = %s, last_update = NOW()WHERE barcode = %s AND creator_id = %s", (new_inventory, barcode, self.creator_id))
                    self.db_manager.execute_query("INSERT INTO inventory_transactions (barcode, product_name, change_quantity, transaction_date, reason, creator_id, new_inventory) VALUES (%s, %s, %s, NOW(), %s,%s, %s)", (barcode, product_name, change_quantity, "冲账", self.creator_id, new_inventory))
                # 标记原始交易为已冲账
                mark_reversed_query = "UPDATE transactions SET reversed = TRUE,remark = %s WHERE id = %s"
                self.db_manager.execute_query(mark_reversed_query, (old_remark+'(已冲账)',transaction_id,))

                # 如果项目名称存在，则更新项目次数
                if project:
                    query = "UPDATE member_beauty_projects SET sessions = sessions + 1 WHERE member_id = %s AND beauty_project_id = (SELECT id FROM beauty_projects WHERE name = %s)"
                    self.db_manager.execute_query(query, (member_id, project))

                # 提交事务
                self.db_manager.commit_transaction()

                # 刷新交易表
                self.search_transaction()

                QMessageBox.information(self, "成功", "交易已成功冲账。")
        except Exception as e:
            # 发生错误，回滚事务
            self.db_manager.rollback_transaction()
            
            QMessageBox.warning(self, "错误", str(e))
            print(e)
    #排序
    def sort_transaction_table(self, column):
        self.transaction_table.sortItems(column, self.sort_order)
        self.sort_order = Qt.DescendingOrder if self.sort_order == Qt.AscendingOrder else Qt.AscendingOrder
    #查询美容项目
    def search_beauty_project(self):
        keyword = self.beauty_project_entry.text()
        # Include the new fields in the query
        query = "SELECT id, name, times, single_price, card_price, duration, DATE_FORMAT(datetime, '%%Y-%%m-%%d %%H:%%i:%%s') FROM beauty_projects WHERE name LIKE %s AND creator_id = %s"
        params = ('%' + keyword + '%', self.creator_id)
        beauty_projects = self.db_manager.execute_query(query, params)
        self.beauty_project_table.setRowCount(0)
        for row, beauty_project in enumerate(beauty_projects):
            self.beauty_project_table.insertRow(row)
            for col, data in enumerate(beauty_project):
                item = QTableWidgetItem(str(data))
                if col == 5:  # 这里，5 表示 duration 所在的列
                    item.setText(f"{data}分钟")
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                self.beauty_project_table.setItem(row, col, item)


    #新建美容项目
    def create_add_beauty_project_window(self):
        add_beauty_project_window = AddBeautyProjectWindow(self)
        add_beauty_project_window.exec_()
    #修改美容项目
    def edit_beauty_project(self):
        selected_row = self.beauty_project_table.currentRow()
        if selected_row != -1:
            beauty_project_id = self.beauty_project_table.item(selected_row, 0).text()
            edit_beauty_project_window = EditBeautyProjectWindow(beauty_project_id, self)
            edit_beauty_project_window.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要修改的美容项目。")
    # 删除美容项目
    def delete_beauty_project(self):
        selected_row = self.beauty_project_table.currentRow()
        if selected_row != -1:
            beauty_project_id = self.beauty_project_table.item(selected_row, 0).text()
            query = f"SELECT * FROM beauty_projects WHERE id = {beauty_project_id}"
            beauty_project = self.db_manager.execute_query(query)[0]
            if beauty_project:
                confirm = QMessageBox.question(self, "确认", "【警告】确定要删除该美容项目吗？")
                if confirm == QMessageBox.Yes:
                    # 检查是否有引用这个美容项目的会员项目
                    check_member_projects_query = f"SELECT * FROM member_beauty_projects WHERE beauty_project_id = {beauty_project_id}"
                    member_projects = self.db_manager.execute_query(check_member_projects_query)
                    if member_projects:
                        QMessageBox.warning(self, "错误", "已产生消费记录，无法删除该美容项目。")
                        return
                    
                    # 如果没有引用，就删除美容项目
                    delete_project_query = f"DELETE FROM beauty_projects WHERE id = {beauty_project_id}"
                    self.db_manager.execute_query(delete_project_query)
                    
                    QMessageBox.information(self, "提示", "美容项目删除成功。")
                    self.search_beauty_project()
            else:
                QMessageBox.warning(self, "错误", "选择的美容项目不存在。")
        else:
            QMessageBox.warning(self, "警告", "请选择要删除的美容项目。")
    #查询营业员
    def search_clerk(self):
        keyword = self.clerks_entry.text()
        query = "SELECT id, name, phone FROM clerks WHERE name LIKE %s AND is_deleted = 0 AND creator_id = %s"
        params = ('%' + keyword + '%', self.creator_id)
        clerks = self.db_manager.execute_query(query, params)
        self.clerks_table.setRowCount(0)
        for row, clerks in enumerate(clerks):
            self.clerks_table.insertRow(row)
            for col, data in enumerate(clerks):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled) #高亮修改这里
                self.clerks_table.setItem(row, col, item)
    #新增营业员窗口
    def create_add_clerk_window(self):
        add_clerk_window = AddClerkWindow(self)
        add_clerk_window.exec_()
    #修改营业员窗口
    def change_clerk_window(self):
        selected_row = self.clerks_table.currentRow()
        if selected_row != -1:
            clerk_id = self.clerks_table.item(selected_row, 0).text()
            change_clerk_window = ChangeClerkWindow(clerk_id,self)
            change_clerk_window.exec_()
        else:
            QMessageBox.warning(self, "警告", "请选择要修改的营业员。")
        
    # 营业员删除
    def delete_clerks(self):
        selected_row = self.clerks_table.currentRow()
        if selected_row != -1:
            clerks_id = self.clerks_table.item(selected_row, 0).text()
            query = f"SELECT * FROM clerks WHERE id = {clerks_id}"
            clerks = self.db_manager.execute_query(query)[0]
            if clerks:
                confirm = QMessageBox.question(self, "确认", "【警告】确定要删除该营业员吗？")
                if confirm == QMessageBox.Yes:
                    # 不真正删除营业员，而是将其标记为已删除
                    query = f"UPDATE clerks SET is_deleted = 1 WHERE id = {clerks_id}"
                    self.db_manager.execute_query(query)
                    QMessageBox.information(self, "提示", "营业员删除成功。")
                    self.search_clerk()
            else:
                QMessageBox.warning(self, "错误", "选择的营业员不存在。")
        else:
            QMessageBox.warning(self, "警告", "请选择要删除的营业员。")
    #打开会员日设置界面
    def show_birthday_rules_dialog(self):
        dialog = BirthdayRulesDialog(self)
        dialog.exec_()
    #查询库存
    def search_inventory(self):
        keyword = self.inventory_entry.text()
        query = "SELECT * FROM product_inventory WHERE (barcode LIKE %s OR product_name LIKE %s) AND creator_id = %s"
        params = ('%' + keyword + '%', '%' + keyword + '%', self.creator_id)
        inventorys = self.db_manager.execute_query(query, params)
        self.inventory_table.setRowCount(0)
        for row, inventorys in enumerate(inventorys):
            self.inventory_table.insertRow(row)
            for col, data in enumerate(inventorys):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled) #高亮修改这里
                self.inventory_table.setItem(row, col, item)
    #打开导入库存对话框
    def open_import_inventory_dialog(self):
        dialog = ImportInventoryDialog(self.creator_id, self)
        dialog.exec_()
    #打开盘点库存对话框
    def open_inventoryCheck_dialog(self):
        dialog = InventoryCheckDialog(self.creator_id, self)
        dialog.exec_()
    
    #注销
    def logout(self):
        # 关闭当前窗口
        self.close()

        # 创建并显示登录窗口
        self.login_window = LoginWindow()
        self.login_window.show()
    #查询会员项目剩余次数
    def search_report(self):
        keyword = self.report_entry.text()
        query = """
            SELECT members.name, members.phone, beauty_projects.name, member_beauty_projects.sessions 
            FROM member_beauty_projects 
            JOIN members ON member_beauty_projects.member_id = members.id AND members.creator_id = %s
            JOIN beauty_projects ON member_beauty_projects.beauty_project_id = beauty_projects.id 
            WHERE (members.name LIKE %s OR members.phone LIKE %s)
        """
        params = (self.creator_id, '%' + keyword + '%', '%' + keyword + '%')
        reports = self.db_manager.execute_query(query, params)

        self.report_table.setRowCount(0)
        for row, report in enumerate(reports):
            self.report_table.insertRow(row)
            for col, data in enumerate(report):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled) #高亮修改这里
                self.report_table.setItem(row, col, item)
    
    def show_business_hours_dialog(self):
        # 创建并显示营业时间设置的对话框
        dialog = BusinessHoursDialog(self)
        dialog.exec_()

    #修改密码
    def show_change_password_dialog(self):
        # 从数据库中检索用户名
        result = self.db_manager.execute_query("SELECT username FROM users WHERE id = %s", (self.creator_id,))[0]


        if result:
            current_username = result[0]
        else:
            QMessageBox.warning(self, "错误", "用户不存在。")
            return

        dialog = ChangePasswordDialog(self, current_username)  # 传入当前用户名
        dialog.exec_()
    #会员等级折扣设置
    def show_discount_settings_dialog(self):
        dialog = DiscountSettingsDialog(self)
        dialog.exec_()
    #会员等级规则设置
    def show_level_rules_dialog(self):
        # 创建并显示会员等级规则对话框
        dialog = LevelRuleSettingsDialog(self)
        dialog.exec_()
    # 美容床位设置对话框
    def show_beauty_bed_dialog(self):
        dialog = BeautyBedDialog(self)
        dialog.exec_()

#导入库存对话框
class ImportInventoryDialog(QDialog):
    def __init__(self,creator_id, parent=None):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle('导入库存')
        self.layout = QVBoxLayout(self)
        self.creator_id = creator_id
        # 说明标签
        self.info_label = QLabel("首先，请确保您的数据符合模板格式。您可以下载模板进行检查。")
        self.layout.addWidget(self.info_label)
        
        # 下载模板按钮
        self.download_template_button = QPushButton("下载模板")
        self.download_template_button.clicked.connect(self.download_template)
        self.layout.addWidget(self.download_template_button)
        
        # 上传文件按钮
        self.upload_file_button = QPushButton("上传库存文件")
        self.upload_file_button.clicked.connect(self.upload_file)
        self.layout.addWidget(self.upload_file_button)

    def download_template(self):
        # 导入库存模板文件下载链接
        template_url = "https://dwpmall.timierhouse.com/updata/temp.xlsx"
        QDesktopServices.openUrl(QUrl(template_url))
        
    def upload_file(self):
        xlsx_file, _ = QFileDialog.getOpenFileName(self, "选择库存数据文件", "", "Excel files (*.xlsx)")
        if xlsx_file:
            failed_rows = []  # 用于记录导入失败的行号
            success_count = 0  # 成功导入的记录数
            failed_count = 0  # 失败的记录数
            try:
                # 打开并读取xlsx文件
                workbook = openpyxl.load_workbook(xlsx_file)
                sheet = workbook.active
                for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # 从第二行开始读取数据，并记录行号
                    barcode = row[0].value  # 条码
                    product_name = row[1].value  # 商品名称
                    import_quantity = row[2].value  # 导入数量

                    # 检查barcode是否为整数类型
                    if not isinstance(barcode, int):
                        failed_rows.append([row_index, barcode, product_name, import_quantity,"条码不是整数类型"])
                        failed_count += 1
                        continue  # 继续处理下一行
                    # 检查import_quantity是否为整数类型
                    if not isinstance(import_quantity, int):
                        failed_rows.append([row_index, barcode, product_name, import_quantity,"导入数量不是整数类型"])
                        failed_count += 1
                        continue  # 继续处理下一行
                    try:
                        query = "SELECT COUNT(*) FROM product_info WHERE barcode = %s AND (creator_id = %s OR creator_id = 2)"
                        params = (barcode, self.creator_id)
                        inforesult = self.db_manager.execute_query(query, params)
                        if inforesult[0][0] == 0:
                            query = """
                                INSERT INTO product_info (barcode, product_name, retail_price, creator_id)
                                VALUES (%s, %s, "0",  %s)
                                """
                            params = (barcode, product_name, self.creator_id)
                            self.db_manager.execute_query(query, params)

                        query = "SELECT inventory_quantity FROM product_inventory WHERE barcode = %s AND creator_id = %s"
                        params = (barcode, self.creator_id)
                        result = self.db_manager.execute_query(query, params)
                        old_inventory_quantity = result[0][0] if result else None

                        if old_inventory_quantity is not None:
                            new_inventory_quantity = old_inventory_quantity + import_quantity
                            query = "UPDATE product_inventory SET inventory_quantity = %s, last_update = NOW() WHERE barcode = %s AND creator_id = %s"
                            params = (new_inventory_quantity, barcode, self.creator_id)
                        else:
                            new_inventory_quantity = import_quantity
                            query = """
                            INSERT INTO product_inventory (barcode, product_name, inventory_quantity, last_update, creator_id)
                            VALUES (%s, %s, %s, NOW(), %s)
                            """
                            params = (barcode, product_name, import_quantity, self.creator_id)
                        self.db_manager.execute_query(query, params)
                        self.db_manager.execute_query("INSERT INTO inventory_transactions (barcode, product_name, change_quantity, transaction_date, reason, creator_id, new_inventory) VALUES (%s, %s, %s, NOW(), %s,%s, %s)", (barcode, product_name,import_quantity, "导入", self.creator_id, new_inventory_quantity))
                        # 假设所有数据库操作成功，则增加成功计数
                        success_count += 1
                    # 在捕获异常的地方记录详细的错误信息
                    except Exception as e:
                        failed_rows.append([row_index, barcode, product_name,import_quantity, str(e)])  # 修改此处以记录更详细的失败信息
                        failed_count += 1
                        continue

                    
                # 保存失败明细到Excel
                if failed_rows:
                    failed_workbook = openpyxl.Workbook()
                    failed_sheet = failed_workbook.active
                    failed_sheet.append(["行号", "条码", "商品名称", "导入数量","错误信息"])
                    for row in failed_rows:
                        failed_sheet.append(row)
                    failed_filename = "导入失败明细.xlsx"
                    failed_workbook.save(failed_filename)
                    response = QMessageBox.critical(self, "导入完成", f"成功导入{success_count}条，失败{failed_count}条。\n失败明细已保存至：{failed_filename}。\n是否现在打开？", QMessageBox.Yes | QMessageBox.No)
                    if response == QMessageBox.Yes:
                        try:
                            os.startfile(failed_filename)
                        except Exception as e:
                            QMessageBox.critical(self, "打开文件失败", f"无法自动打开文件：{failed_filename}。\n请手动打开。\n错误详情：{e}")
                else:
                    QMessageBox.information(self, "导入成功", f"全部数据成功导入！共导入{success_count}条记录。")
                # 移动文件到"已导入"文件夹
                imported_folder = os.path.join(os.path.dirname(xlsx_file), "已导入")
                if not os.path.exists(imported_folder):
                    os.makedirs(imported_folder)
                shutil.move(xlsx_file, os.path.join(imported_folder, os.path.basename(xlsx_file)))
                QMessageBox.information(self, "文件已移动", f"原文件已移动到：{imported_folder}")

                self.close()
            except Exception as e:
                QMessageBox.critical(self, "导入失败", f"导入过程中发生错误：{e}")

#盘点库存对话框
class InventoryCheckDialog(QDialog):
    def __init__(self,creator_id, parent=None):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle('盘点库存')
        self.layout = QVBoxLayout(self)
        self.creator_id = creator_id
        # 说明标签
        self.info_label = QLabel("请下载盘点单数据文件。确保您的数据格式正确。")
        self.layout.addWidget(self.info_label)
        
        # 下载模板按钮
        self.download_template_button = QPushButton("下载盘点单")
        self.download_template_button.clicked.connect(self.export_InventoryCheck)
        self.layout.addWidget(self.download_template_button)
        
        # 上传文件按钮
        self.upload_file_button = QPushButton("上传盘点单")
        self.upload_file_button.clicked.connect(self.upload_file)
        self.layout.addWidget(self.upload_file_button)

    def export_InventoryCheck(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 添加表头
        headers = ["条码", "商品名称", "零售单价" ,"库存数量", "实盘数量"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # 添加数据
        query = """
        SELECT pi.barcode, pi.product_name, pi.retail_price, inv.inventory_quantity
        FROM product_inventory inv
        JOIN product_info pi ON inv.barcode = pi.barcode
        WHERE inv.creator_id = %s
        """
        params = (self.creator_id,)
        InventoryChecks = self.db_manager.execute_query(query, params)

        for row_num, InventoryCheck in enumerate(InventoryChecks, 2):
            for col_num, value in enumerate(InventoryCheck, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # 选择保存路径
        filename, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel Files (*.xlsx)")
        if filename:
            workbook.save(filename)
            QMessageBox.information(self, "提示", f"盘点单已成功下载：{filename}")
        
    def upload_file(self):
        xlsx_file, _ = QFileDialog.getOpenFileName(self, "选择盘点数据文件", "", "Excel files (*.xlsx)")
        if xlsx_file:
            failed_rows = []  # 用于记录导入失败的行号
            try:
                # 打开并读取xlsx文件
                workbook = openpyxl.load_workbook(xlsx_file)
                sheet = workbook.active
                for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # 从第二行开始读取数据，并记录行号
                    try:
                        barcode = row[0].value  # 条码
                        product_name = row[1].value  # 商品名称
                        inventory_quantity = row[3].value  # 库存数量
                        check_inventory_quantity = row[4].value #实盘数量
                        #盘点差异
                        inventory_discrepancy = int(check_inventory_quantity) - int(inventory_quantity)
                        #检查是否存在商品信息
                        query = "SELECT COUNT(*) FROM product_info WHERE barcode = %s AND (creator_id = %s OR creator_id = 2)"
                        params = (barcode, self.creator_id)
                        inforesult = self.db_manager.execute_query(query, params)
                        if inforesult[0][0] == 0:
                            query = """
                                INSERT INTO product_info (barcode, product_name, retail_price, creator_id)
                                VALUES (%s, %s, "0",  %s)
                                """
                            params = (barcode, product_name, self.creator_id)
                            self.db_manager.execute_query(query, params)

                        #更新库存
                        query = "UPDATE product_inventory SET inventory_quantity = %s, last_update = NOW() WHERE barcode = %s AND creator_id = %s"
                        params = (check_inventory_quantity, barcode, self.creator_id)
                        self.db_manager.execute_query(query, params)
                        #创建流水记录
                        self.db_manager.execute_query("INSERT INTO inventory_transactions (barcode, product_name, change_quantity, transaction_date, reason, creator_id, new_inventory) VALUES (%s, %s, %s, NOW(), %s,%s, %s)", (barcode, product_name,inventory_discrepancy, "盘点", self.creator_id, check_inventory_quantity))
                    # 在捕获异常的地方记录详细的错误信息
                    except Exception as e:
                        failed_rows.append(f"行{row_index}: 条码{barcode}，商品名称{product_name}（错误：{e}）")
                        continue  # 继续处理下一行

                    
                if failed_rows:
                    QMessageBox.critical(self, "导入完成", f"导入过程中以下行出现错误并被跳过：{', '.join(map(str, failed_rows))}")
                else:
                    QMessageBox.information(self, "导入成功", "盘点数据已成功导入!")
                self.close()
            except Exception as e:
                QMessageBox.critical(self, "导入失败", f"导入过程中发生错误：{e}")
#单品流水对话框
class QueryProductFlowDialog(QDialog):
    def __init__(self, barcode, creator_id,parent):  # 修正参数顺序和self的使用
        super().__init__(parent)
        self.setWindowTitle("单品库存流水")
        self.resize(1000, 600)
        self.db_manager = DatabaseManager()
        self.barcode = barcode
        self.creator_id = creator_id

        # 布局和表格设置
        layout = QVBoxLayout()
        self.inventory_transactions = QTableWidget()
        self.inventory_transactions.setColumnCount(8)
        self.inventory_transactions.setHorizontalHeaderLabels(["ID", "商品条码", "商品名称", "发生数量", "发生时间", "变动原因", "门店ID", "数量平衡"])
        self.inventory_transactions.setSelectionBehavior(QTableWidget.SelectRows)
        self.inventory_transactions.setColumnHidden(0, True)
        self.inventory_transactions.setColumnHidden(6, True)
        self.inventory_transactions.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        layout.addWidget(self.inventory_transactions)

        # 添加下载按钮
        self.download_button = QPushButton("下载导出")
        self.download_button.clicked.connect(self.download_as_excel)
        layout.addWidget(self.download_button)

        self.setLayout(layout)
        self.check_inventory_transactions()

    def check_inventory_transactions(self):
        query = """
            SELECT *
            FROM inventory_transactions
            WHERE barcode = %s AND creator_id =%s
        """
        params = (self.barcode,self.creator_id)
        reports = self.db_manager.execute_query(query, params)

        self.inventory_transactions.setRowCount(0)
        for row, report in enumerate(reports):
            self.inventory_transactions.insertRow(row)
            for col, data in enumerate(report):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled) #高亮修改这里
                self.inventory_transactions.setItem(row, col, item)
    def download_as_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 添加表头
        headers = ["条码", "商品名称", "发生数量" ,"发生时间", "变动原因","数量平衡"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # 添加数据
        query = """
            SELECT barcode,product_name,change_quantity,transaction_date,reason,new_inventory
            FROM inventory_transactions
            WHERE barcode = %s AND creator_id =%s
            """
        params = (self.barcode,self.creator_id)
        reports = self.db_manager.execute_query(query, params)

        for row_num, report in enumerate(reports, 2):
            for col_num, value in enumerate(report, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # 选择保存路径
        filename, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel Files (*.xlsx)")
        if filename:
            workbook.save(filename)
            QMessageBox.information(self, "提示", f"已成功下载：{filename}")



#新增会员对话框
class AddMemberWindow(QDialog):
    def __init__(self,creator_id,parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.creator_id = creator_id  # 保存 creator_id 到实例变量中
        
        self.setWindowTitle("新增会员")
        name_label = QLabel("会员姓名：")
        self.name_entry = QLineEdit()
        phone_label = QLabel("手机号：")
        self.phone_entry = QLineEdit()
        birthday_label = QLabel("生日：")  # 新增
        self.birthday_entry = QDateEdit()  # 新增
        self.birthday_entry.setCalendarPopup(True)  # 新增

        layout = QGridLayout()
        layout.addWidget(name_label, 0, 0)
        layout.addWidget(self.name_entry, 0, 1)
        layout.addWidget(phone_label, 1, 0)
        layout.addWidget(self.phone_entry, 1, 1)
        layout.addWidget(birthday_label, 2, 0)  # 新增
        layout.addWidget(self.birthday_entry, 2, 1)  # 新增

        add_button = QPushButton("确定")
        add_button.clicked.connect(self.add_member)

        layout.addWidget(add_button, 3, 0, 1, 2)  # 修改

        self.setLayout(layout)

    def validate_phone(self, phone):
        pattern = r"^1[3456789]\d{9}$"
        if not re.match(pattern, phone):
            return False
        return True
    
    def check_phone_exists(self, phone):
        query = f"SELECT * FROM members WHERE phone = '{phone}'"
        member = self.db_manager.execute_query(query)
        if member:
            return True
        return False

    def add_member(self):
        name = self.name_entry.text()
        phone = self.phone_entry.text()
        birthday = self.birthday_entry.date().toString("yyyy-MM-dd")
        if not self.validate_phone(phone):
            QMessageBox.warning(self, "错误", "手机号格式不正确。")
            return
        if self.check_phone_exists(phone):
            QMessageBox.warning(self, "错误", "手机号已存在，请输入其他手机号。")
            return
        query = f"INSERT INTO members (name, phone, birthday, register_date, balance, creator_id) VALUES ('{name}', '{phone}', '{birthday}', NOW(), 0, {self.creator_id})"  # 修改
        self.db_manager.execute_query(query)
        QMessageBox.information(self, "提示", "会员添加成功。")
        self.parent().search_member()
        self.close()


#修改会员信息
class EditMemberWindow(QDialog):
    def __init__(self, member_id, parent):
        super().__init__(parent)
        self.setWindowTitle("修改会员信息")
        self.db_manager = DatabaseManager()
        self.member_id = member_id

        query = f"SELECT * FROM members WHERE id = {self.member_id}"
        member = self.db_manager.execute_query(query)[0]
        self.old_balance = member[6]

        name_label = QLabel("会员姓名：")
        self.name_entry = QLineEdit()
        self.name_entry.setText(member[1])
        phone_label = QLabel("手机号：")
        self.phone_entry = QLineEdit()
        self.phone_entry.setText(member[2])
        birthday_label = QLabel("生日：")
        self.birthday_entry = QDateEdit()
        self.birthday_entry.setCalendarPopup(True)
        self.birthday_entry.setDate(QDate.fromString(member[5].strftime("%Y-%m-%d"), "yyyy-MM-dd"))  
        balance_label = QLabel("余额：")
        self.balance_entry = QLineEdit()
        self.balance_entry.setText(str(member[6]))
        level_label = QLabel("会员等级：")
        self.level_combo = QComboBox()
        self.level_combo.addItems(["普通卡", "银卡", "金卡", "钻卡", "至尊卡"])
        self.level_combo.setCurrentText(member[9])  # 注意: 这里的索引9可能需要更改，具体取决于`membership_level`字段在数据库查询结果中的位置



        layout = QGridLayout()
        layout.addWidget(name_label, 0, 0)
        layout.addWidget(self.name_entry, 0, 1)
        layout.addWidget(phone_label, 1, 0)
        layout.addWidget(self.phone_entry, 1, 1)
        layout.addWidget(birthday_label, 2, 0)
        layout.addWidget(self.birthday_entry, 2, 1)
        layout.addWidget(balance_label, 3, 0)
        layout.addWidget(self.balance_entry, 3, 1)
        layout.addWidget(level_label, 4, 0)
        layout.addWidget(self.level_combo, 4, 1)


        save_button = QPushButton("保存")
        save_button.clicked.connect(self.save_member)

        layout.addWidget(save_button, 5, 0, 1, 2)

        self.setLayout(layout)
    def validate_phone(self, phone):
        pattern = r"^1[3456789]\d{9}$"
        if not re.match(pattern, phone):
            return False
        return True
    
    def check_phone_exists(self, phone):
        query = f"SELECT * FROM members WHERE phone = '{phone}'"
        member = self.db_manager.execute_query(query)
        if member:
            return True
        return False

    def save_member(self):
        name = self.name_entry.text()
        phone = self.phone_entry.text()
        birthday = self.birthday_entry.date().toString("yyyy-MM-dd")
        balance = self.balance_entry.text()
        membership_level = self.level_combo.currentText()

        # 检查手机号是否已经被修改
        query = f"SELECT phone, membership_level FROM members WHERE id = {self.member_id}"
        result = self.db_manager.execute_query(query)[0]
        old_phone = result[0]
        old_level = result[1]

        if phone != old_phone:
            QMessageBox.warning(self, "错误", "手机号已存在，请输入其他手机号。")
            return
        if not self.validate_phone(phone):
            QMessageBox.warning(self, "错误", "手机号格式不正确。")
            return

        # 更新会员信息
        query = f"UPDATE members SET name = '{name}', phone = '{phone}', birthday = '{birthday}', balance = {balance}, membership_level = '{membership_level}', creator_id = {self.parent().creator_id} WHERE id = {self.member_id}"
        self.db_manager.execute_query(query)

        # 检查余额是否发生了变化，如果是，则记录交易
        query = f"SELECT balance, membership_level FROM members WHERE id = {self.member_id}"
        new_balance = self.db_manager.execute_query(query)[0][0]
        membership_level = self.db_manager.execute_query(query)[0][1]
        if self.old_balance != new_balance:
            amount = float(new_balance) - float(self.old_balance)
            transaction_query = f"INSERT INTO transactions (member_id, amount, project,transaction_date, balance, remark, membership_level, phone, pay_method) VALUES (%s, %s, '无',NOW(), %s,'修改余额', %s, %s,'未知')"
            params = (self.member_id, amount, new_balance, membership_level,phone)
            inserted_id  = self.db_manager.execute_insert(transaction_query,params)

        # 如果会员等级发生了变化
        if membership_level != old_level:
            # 从membership_level_rules表中获取新等级的upgrade_amount
            level_rule_query = f"SELECT upgrade_amount FROM membership_level_rules WHERE membership_level = '{membership_level}'"
            upgrade_amount = self.db_manager.execute_query(level_rule_query)
            if upgrade_amount:
                upgrade_amount_value = upgrade_amount[0][0]
                # 更新会员的total_recharge_amount为新等级的upgrade_amount
                update_recharge_amount_query = f"UPDATE members SET total_recharge_amount = {upgrade_amount_value} WHERE id = {self.member_id}"
                self.db_manager.execute_query(update_recharge_amount_query)

        QMessageBox.information(self, "提示", "会员信息修改成功。")
        self.parent().search_transaction()
        self.parent().search_member()
        self.close()


#会员充值
class RechargeDialog(QDialog):
    def __init__(self, member_id, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("会员充值")

        self.member_id = member_id
        self.creator_id = parent.creator_id
        # 查询当前会员的名字和余额
        query = f"SELECT name, balance FROM members WHERE id = {self.member_id}"
        member_info = self.db_manager.execute_query(query)[0]
        member_name = member_info[0]
        member_balance = member_info[1]

        # 创建标签以显示会员的名字和余额
        member_info_label = QLabel(f"会员姓名：{member_name}\n会员余额：{member_balance}元")
        member_info_label.setStyleSheet("font-weight: bold;")

        amount_label = QLabel("充值金额：")
        self.amount_entry = QLineEdit()
        
        remark_label= QLabel("备注：")
        self.remark_entry = QLineEdit()

        clerk_label = QLabel("营业员：")
        self.clerk_combo = QComboBox()
        self.load_clerks()
        self.resize(300, 200)

        layout = QVBoxLayout()  # 定义布局
        layout.addWidget(member_info_label)
        layout.addWidget(amount_label)
        layout.addWidget(self.amount_entry)
        layout.addWidget(remark_label)
        layout.addWidget(self.remark_entry)
        layout.addWidget(clerk_label)
        layout.addWidget(self.clerk_combo)

        confirm_button = QPushButton("确认充值")
        confirm_button.clicked.connect(self.confirm_recharge)

        layout.addWidget(confirm_button)

        self.setLayout(layout)
    #读取营业员信息
    def load_clerks(self):
        query = f"SELECT name FROM clerks WHERE creator_id = {self.creator_id}"
        clerks = self.db_manager.execute_query(query)
        for clerk in clerks:
            self.clerk_combo.addItem(clerk[0])
    #确认充值信息
    def confirm_recharge(self):
        try:
            amount = float(self.amount_entry.text())
            remark = self.remark_entry.text()
            remark = remark if remark else ""
            if amount > 0:
                query = f"SELECT balance, total_recharge_amount, membership_level,phone FROM members WHERE id = {self.member_id}"
                member = self.db_manager.execute_query(query)[0]
                membership_level = self.db_manager.execute_query(query)[0][2]
                phone = self.db_manager.execute_query(query)[0][3]
                if member:
                    current_balance = member[0]
                    current_total_recharge_amount = member[1]
                    current_level = member[2]
                    new_balance = current_balance + amount
                    new_total_recharge_amount = float(current_total_recharge_amount) + amount

                    clerk_name = self.clerk_combo.currentText()
                    clerk = self.db_manager.execute_query("SELECT id FROM clerks WHERE name = %s", (clerk_name,))
                    if clerk:
                        clerk_id = clerk[0]

                        query = f"UPDATE members SET balance = {new_balance}, total_recharge_amount = {new_total_recharge_amount} WHERE id = {self.member_id}"
                        self.db_manager.execute_query(query)

                        transaction_query = "INSERT INTO transactions (member_id, amount, project,transaction_date, balance, remark, clerk_id, membership_level,phone) VALUES (%s, %s, %s, NOW(), %s, %s, %s, %s, %s)"
                        params = (self.member_id, amount, '无', new_balance, '充值|'+ remark, clerk_id, membership_level,phone)
                        inserted_id  = self.db_manager.execute_insert(transaction_query, params)
                        # 查询所有的会员等级及其对应的金额要求
                        query = "SELECT membership_level, upgrade_amount FROM membership_level_rules ORDER BY upgrade_amount DESC"
                        levels = self.db_manager.execute_query(query)

                        # 确定新的会员等级
                        new_level = None
                        for level, amount_required in levels:
                            if new_total_recharge_amount >= amount_required:
                                new_level = level
                                break

                        # 如果新的会员等级与当前会员等级不同，并且新的等级比当前等级高
                        if new_level and new_level != current_level and LEVELS.index(new_level) > LEVELS.index(current_level):
                            # 更新会员等级并记录升级时间
                            query = f"UPDATE members SET membership_level = '{new_level}', membership_level_last_updated = NOW() WHERE id = {self.member_id}"
                            self.db_manager.execute_query(query)
                            QMessageBox.information(self, "提示", f"恭喜！会员等级已升级到 {new_level}。")

                        QMessageBox.information(self, "提示", "充值成功。")
                        self.parent().search_member()
                        self.parent().search_transaction()
                        self.close()
                    else:
                        QMessageBox.information(self, "提示", "请输入营业员！")
                        return
                            
                else:
                    QMessageBox.warning(self, "错误", "选择的会员不存在。")
            else:
                QMessageBox.warning(self, "错误", "充值金额必须大于0。")
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的充值金额。")

#会员购买商品窗口
class PurchaseProductDialog(QDialog):
    def __init__(self, member_id, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.member_id = member_id
        self.creator_id = parent.creator_id
        self.setWindowTitle("购买商品")
        

        # 查询当前会员的名字和余额
        query = "SELECT name, balance,membership_level FROM members WHERE id = %s"
        member_info = self.db_manager.execute_query(query, (self.member_id,))[0]
        member_name = member_info[0]
        member_balance = member_info[1]
        membership_level = member_info[2]

        # 创建标签以显示会员的名字和余额
        member_info_label = QLabel(f"会员姓名：{member_name}\n会员余额：{member_balance}元\n会员等级：{membership_level}")
        member_info_label.setStyleSheet("font-weight: bold;color: red")

        search_label = QLabel("购买商品：")
        search_label.setStyleSheet("font-weight: bold;")
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("输入条码或名称")

        self.search_entry.textChanged.connect(self.delayed_search)
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.update_products_based_on_search)

        product_label = QLabel("商品选择：")
        self.product_combo = QComboBox()
        # 创建商品选择表格
        self.products_table = QTableWidget(0, 6)  # 初始无行，4列
        self.products_table.setHorizontalHeaderLabels(['商品条码','商品名称', '零售价', '数量', '小计', '折扣'])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        # 设置数量列为可编辑
        self.products_table.setEditTriggers(QTableWidget.AllEditTriggers)
        # 连接cellChanged信号
        self.products_table.cellChanged.connect(self.on_cell_changed)


        # 添加商品按钮
        add_product_button = QPushButton("添加商品到购物车")
        add_product_button.clicked.connect(self.add_product_to_cart)
        # 创建删除商品按钮
        delete_product_button = QPushButton("删除行")
        delete_product_button.clicked.connect(self.delete_selected_product_from_cart)
        


        amount_label = QLabel("实付金额：")
        self.amount_entry = QLineEdit("0")
        self.amount_entry.setReadOnly(True)  # 设置为只读，因为金额由选择的商品决定
        # 创建结算方式标签和下拉框
        payment_method_label = QLabel("结算方式：")
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItems(["账户余额", "现付"])

        clerk_label = QLabel("营业员：")
        self.clerk_combo = QComboBox()

        self.load_products()

        # 创建确认购买按钮
        confirm_button = QPushButton("确认购买")
        confirm_button.clicked.connect(self.confirm_purchase)

        # 创建布局
        layout = QVBoxLayout()
        layout.addWidget(member_info_label)
        layout.addWidget(search_label)
        layout.addWidget(self.search_entry)
        layout.addWidget(self.product_combo)
        layout.addWidget(add_product_button)
        layout.addWidget(product_label)
        layout.addWidget(self.products_table)
        layout.addWidget(delete_product_button)
        # 创建一个水平布局，用于放置实付金额和结算方式
        amount_layout = QHBoxLayout()
        amount_layout.addWidget(amount_label)
        amount_layout.addWidget(self.amount_entry)
        amount_layout.addWidget(payment_method_label)
        amount_layout.addWidget(self.payment_method_combo)
        layout.addLayout(amount_layout)

        layout.addWidget(clerk_label)
        layout.addWidget(self.clerk_combo)
        layout.addWidget(confirm_button)

        self.setLayout(layout)
        # 设置窗口大小
        self.resize(800, 700)
    
    def set_table_item_style(self, item, color='black'):
        item.setForeground(QBrush(QColor(color)))
    
    def delete_selected_product_from_cart(self):
        selected_rows = self.products_table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先选择要删除的商品行")
            return
        
        for selected_row in selected_rows:
            self.products_table.removeRow(selected_row.row())
        
        self.update_total_amount()  # 删除行后，更新总金额
    def add_product_to_cart(self):
        try:
            self.products_table.cellChanged.disconnect(self.on_cell_changed)
        except TypeError:
            pass
        # 获取当前选中的商品信息
        current_index = self.product_combo.currentIndex()
        if current_index == -1:
            QMessageBox.warning(self, "错误", "未选择商品")
            return
        
        # 获取商品价格
        retail_price = self.product_combo.itemData(current_index)[1]
        # 已经存储了商品名称和条形码等其他信息作为显示文本
        current_text = self.product_combo.currentText()
        product_info = current_text.split(" | ")  # 根据你的格式进行调整
        barcode =  product_info[0]
        product_name = product_info[1].split(" - ")[0]
        discount = self.get_member_info_and_discount()[0]
        discount_price = round(float(retail_price) * discount,2)
        # 清除search_entry的内容并让光标停留在search_entry上
        self.search_entry.clear()
        self.search_entry.setFocus()
        
        # 检查表格中是否已有该商品
        product_exists = False
        for row in range(self.products_table.rowCount()):
            if self.products_table.item(row, 1).text() == product_name:
                # 如果找到了，更新数量和小计
                qty_item = self.products_table.item(row, 3)
                qty = int(qty_item.text()) + 1
                qty_item.setText(str(qty))
                
                subtotal_item = self.products_table.item(row, 4)
                subtotal = round(float(discount_price) * qty,2)
                subtotal_item.setText(str(subtotal))
                self.set_table_item_style(subtotal_item, 'red' if discount < 1 else 'black')
                product_exists = True
                break
        
        # 如果表格中没有该商品，则添加新行
        if not product_exists:
            row_position = self.products_table.rowCount()
            self.products_table.insertRow(row_position)
            self.products_table.setItem(row_position, 0, QTableWidgetItem(barcode))
            self.products_table.setItem(row_position, 1, QTableWidgetItem(product_name))
            self.products_table.setItem(row_position, 2, QTableWidgetItem(str(retail_price)))
            self.products_table.setItem(row_position, 3, QTableWidgetItem("1"))  # 默认数量为1
            discount_price = float(discount_price)  # 小计初始化为单价，因为默认数量为1
            self.products_table.setItem(row_position, 4, QTableWidgetItem(str(discount_price)))
            self.set_table_item_style(self.products_table.item(row_position, 4), 'red' if discount < 1 else 'black')
            if discount == 1:
                self.products_table.setItem(row_position, 5, QTableWidgetItem('无折扣'))  # 获取会员折扣
            else:
                self.products_table.setItem(row_position, 5, QTableWidgetItem(str(discount*10)+'折'))  # 获取会员折扣
        # 更新总计
        self.update_total_amount()
        self.products_table.cellChanged.connect(self.on_cell_changed)
    def on_cell_changed(self, row, column):
        # 确保更改的是数量列，这里假设数量列的索引是2
        if column in [2, 3]:
            try:
                # 获取数量和零售价
                quantity = int(self.products_table.item(row, 3).text())
                retail_price = float(self.products_table.item(row, 2).text())
                
                discount = self.get_member_info_and_discount()[0]
                
                # 更新小计，如果应用了折扣则显示红色字体
                subtotal = round(quantity * retail_price * discount, 2)
                subtotal_item = QTableWidgetItem(str(subtotal))
                if discount < 1:
                    subtotal_item.setForeground(QBrush(QColor(255, 0, 0))) # 设置为红色字体
                self.products_table.setItem(row, 4, subtotal_item)
                
                # 更新总金额
                self.update_total_amount()
            except ValueError:
                # 如果发生了ValueError异常，显示警告消息
                QMessageBox.warning(self, "错误", "请输入有效的零售价和数量")
                # 重置零售价和数量为安全值，比如零售价为0，数量为1
                # 注意这里需要断开和重新连接cellChanged信号，以避免无限循环
                self.products_table.cellChanged.disconnect(self.on_cell_changed)
                if column == 1:
                    self.products_table.setItem(row, column, QTableWidgetItem("0"))
                else:
                    self.products_table.setItem(row, column, QTableWidgetItem("1"))
                self.products_table.cellChanged.connect(self.on_cell_changed)

    def load_products(self):
        self.update_products_based_on_search()
        self.load_clerks()
        self.product_combo.setCurrentIndex(-1)  #清除选择
    def delayed_search(self):
        # 设置延迟时间，例如300毫秒
        self.search_timer.start(300)
    def get_member_info_and_discount(self):
        # 获取会员信息
        query = f"SELECT membership_level, birthday FROM members WHERE id = {self.member_id}"
        member_info = self.db_manager.execute_query(query)[0]
        member_level, member_birthday = member_info
        # 检查今天是否为会员日或会员的生日
        discount_info = self.calculate_discount(member_level, member_birthday)
        return discount_info,member_level

    def calculate_discount(self, member_level, member_birthday):
        current_day = datetime.now().day
        current_month = datetime.now().month
        is_member_birthday = (member_birthday.day == current_day) and (member_birthday.month == current_month)
        
        # 获取会员日信息和折扣
        member_day_info = self.db_manager.execute_query("SELECT day_of_month FROM member_birthday_rules WHERE creator_id = %s", (self.creator_id,))
        member_day_discount_info = self.db_manager.execute_query("SELECT member_day_discount, birthday_discount FROM membership_discounts WHERE membership_level = %s AND creator_id = %s", (member_level, self.creator_id))
        if member_day_info:
            is_member_day = member_day_info[0][0] == current_day
        else:
            is_member_day = False
        
        if member_day_discount_info:
            member_day_discount, birthday_discount = member_day_discount_info[0]
        else:
            member_day_discount = birthday_discount = 10  # 默认值
        
        discount = 1  # 默认没有折扣
        if is_member_day:
            discount = member_day_discount / 10
        elif is_member_birthday:
            discount = birthday_discount / 10
        else:
            # 如果不是会员日或生日，则查询会员等级的折扣
            discount_level = self.db_manager.execute_query("SELECT discount FROM membership_discounts WHERE membership_level = %s AND creator_id = %s", (member_level, self.creator_id))
            if discount_level:
                discount = discount_level[0][0] / 10
        
        return discount


    #读取营业员信息
    def load_clerks(self):
        query = f"SELECT name FROM clerks WHERE creator_id = {self.creator_id}"
        clerks = self.db_manager.execute_query(query)
        for clerk in clerks:
            self.clerk_combo.addItem(clerk[0])

    def update_products_based_on_search(self):
        # 停止计时器，以防万一
        self.search_timer.stop()
        creator_id =self.creator_id
        search_text = self.search_entry.text()
        if search_text:
            query = """
            SELECT id, barcode, product_name, retail_price
            FROM product_info
            WHERE (barcode LIKE %s OR product_name LIKE %s)
            AND (creator_id = 2 OR creator_id = %s)
            """
            search_pattern = f'%{search_text}%'
            products = self.db_manager.execute_query(query, (search_pattern, search_pattern, creator_id))
            self.product_combo.clear()  # 清除旧的选项
            for product_id, barcode, product_name, retail_price in products:
                # 更新下拉框
                self.product_combo.addItem(f"{barcode} | {product_name} - {retail_price}元", userData=(product_id, retail_price))
        else:
            # 清除下拉框选择
            self.product_combo.clear()  # 清除旧的选项
    def update_total_amount(self):
        total_amount = 0.0
        for row in range(self.products_table.rowCount()):
            price = float(self.products_table.item(row, 2).text())
            quantity = int(self.products_table.item(row, 3).text())
            subtotal = price * quantity
            total_amount += subtotal

        # 使用提取的方法获取会员信息和折扣
        discount = self.get_member_info_and_discount()[0]
        # 应用折扣
        if discount and discount < 1:
            discounted_price = round(float(total_amount) * discount, 2)
            self.amount_entry.setText(str(discounted_price))
            self.amount_entry.setStyleSheet("color: red;")  # 有折扣，设置字体为红色
        else:
            self.amount_entry.setText(str(total_amount))  # 没有折扣，显示原价
            self.amount_entry.setStyleSheet("color: black;")  # 设置为默认颜色


    #确认购买
    def confirm_purchase(self):
        # 初始化总金额
        total_amount = 0.0
        
        # 获取营业员的信息，因为它在整个购买过程中保持不变
        clerk_name = self.clerk_combo.currentText()
        clerk_query_result = self.db_manager.execute_query("SELECT id FROM clerks WHERE name = %s", (clerk_name,))
        if not clerk_query_result:
            QMessageBox.warning(self, "错误", "请输入有效的营业员！")
            return
        clerk_id = clerk_query_result[0][0]

        # 获取会员的折扣信息
        discount, member_level = self.get_member_info_and_discount()

        # 遍历表格中的每一行，处理每个购买的商品
        for row in range(self.products_table.rowCount()):
            barcode =self.products_table.item(row, 0).text()
            product_name = self.products_table.item(row, 1).text()
            product_price = float(self.products_table.item(row, 2).text())
            quantity = int(self.products_table.item(row, 3).text())
            
            # 计算小计并加到总金额
            subtotal =round(product_price * quantity * discount,2)
            total_amount += subtotal

        # 获取会员信息
        member_info = self.db_manager.execute_query("SELECT phone, balance FROM members WHERE id = %s", (self.member_id,))[0]
        phone, current_balance = member_info
        total_amount =round(total_amount,2)
        # 获取选择的结算方式
        payment_method = self.payment_method_combo.currentText()
        if payment_method == "账户余额":
            # 计算新的余额
            new_balance = round(current_balance - total_amount,2)
            if new_balance < 0:
                QMessageBox.warning(self, "错误", "余额不足，无法完成消费。")
                return

            # 更新会员余额
            self.db_manager.execute_query("UPDATE members SET balance = %s WHERE id = %s", (new_balance, self.member_id))
            QMessageBox.information(self, "购买成功", f"消费金额：{total_amount}\n当前余额：{new_balance}")
        else:
            QMessageBox.information(self, "购买成功", f"消费金额：{total_amount}")
        # 记录每个商品的消费事务
        for row in range(self.products_table.rowCount()):
            barcode =self.products_table.item(row, 0).text()
            product_name = self.products_table.item(row, 1).text()
            product_price = float(self.products_table.item(row, 2).text())
            quantity = int(self.products_table.item(row, 3).text())
            project_value = f"{barcode}|{product_name}*{quantity}"
            subtotal = product_price * quantity * discount
            new_balance = current_balance - subtotal
            if payment_method == "账户余额":
                remark = '购买'
                balance = new_balance
            else:
                remark= '购买现付'
                balance = None
            inserted_id  = self.db_manager.execute_insert("INSERT INTO transactions (member_id, clerk_id, amount, project, transaction_date, balance, membership_level, remark, phone,pay_method) VALUES (%s, %s, %s, %s, NOW(), %s, %s, %s, %s, %s)", (self.member_id, clerk_id, -subtotal, project_value, balance, member_level, remark , phone,payment_method))

            old_inventory = self.db_manager.execute_query("SELECT inventory_quantity FROM product_inventory WHERE creator_id = %s AND barcode = %s", (self.creator_id,barcode))
            if old_inventory:
                new_inventory = old_inventory[0][0] - quantity
                self.db_manager.execute_query("UPDATE product_inventory SET inventory_quantity = %s, last_update = NOW()WHERE barcode = %s AND creator_id = %s", (new_inventory, barcode, self.creator_id))
            else:
                new_inventory = 0
            self.db_manager.execute_query("INSERT INTO inventory_transactions (barcode, product_name, change_quantity, transaction_date, reason, creator_id, new_inventory) VALUES (%s, %s, %s, NOW(), %s,%s, %s)", (barcode, product_name, -quantity, "零售", self.creator_id, new_inventory))
            
            
            current_balance = new_balance
            self.accept()  # 关闭对话框
            # 添加延时
            time.sleep(1)

#会员消费窗口
class ConsumeDialog(QDialog):
    def __init__(self, member_id, appointment_id, parent, beauty_project_id=None, from_complete_service=False):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.actual_amount = 0
        self.beauty_project_id = beauty_project_id
        self.from_complete_service = from_complete_service
        self.appointment_id = appointment_id
        self.clerk_id = None
        self.setWindowTitle("会员消费")

        self.member_id = member_id
        self.creator_id = parent.creator_id
        # 查询当前会员的名字和余额
        query = f"SELECT name, balance FROM members WHERE id = {self.member_id}"
        member_info = self.db_manager.execute_query(query)[0]
        member_name = member_info[0]
        member_balance = member_info[1]

        # 创建标签以显示会员的名字和余额
        member_info_label = QLabel(f"会员姓名：{member_name}\n会员余额：{member_balance}元")
        member_info_label.setStyleSheet("font-weight: bold;")

        amount_label = QLabel("消费金额：")
        self.amount_entry = QLineEdit("0")

        project_label = QLabel("消费项目：")
        self.project_combo = QComboBox()
        
        project_session_label = QLabel("消耗次数：")
        self.project_session_entry = QLineEdit("1")
        self.project_session_entry.returnPressed.connect(self.update_amount)
        clerk_label = QLabel("营业员：")
        self.clerk_combo = QComboBox()
        payment_method_label = QLabel("结算方式：")
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItems(["账户余额", "现付"])
        self.load_clerks()
        self.resize(650, 200)

        # 在 project_combo 初始化完成后连接信号和槽。
        self.project_combo.currentIndexChanged.connect(self.update_amount)
        
        self.load_projects()
        # 创建会员美容项目表格
        self.projects_table = QTableWidget()
        self.projects_table.setColumnCount(2)
        self.projects_table.setHorizontalHeaderLabels(["美容项目", "剩余次数"])
        self.projects_table.verticalHeader().setVisible(False)
        self.load_member_projects()

        # 创建主要布局
        layout = QVBoxLayout()
        layout.addWidget(member_info_label)
        # 创建水平布局
        hbox = QHBoxLayout()

        # 左侧布局
        left_layout = QVBoxLayout()
        left_layout.addWidget(project_label)
        left_layout.addWidget(self.project_combo)
        left_layout.addWidget(project_session_label)
        left_layout.addWidget(self.project_session_entry)
        left_layout.addWidget(amount_label)
        left_layout.addWidget(self.amount_entry)
        left_layout.addWidget(clerk_label)
        left_layout.addWidget(self.clerk_combo)
        left_layout.addWidget(payment_method_label)
        left_layout.addWidget(self.payment_method_combo)

        hbox.addLayout(left_layout)
        # 右侧布局
        right_layout = QVBoxLayout()
        purchased_label = QLabel("已购买项目")
        right_layout.addWidget(purchased_label)
        right_layout.addWidget(self.projects_table)
        hbox.addLayout(right_layout)

        layout.addLayout(hbox)

        confirm_button = QPushButton("确认消费")
        confirm_button.clicked.connect(self.confirm_consume)

        layout.addWidget(confirm_button)
        self.project_session_entry.editingFinished.connect(self.update_amount)
        self.setLayout(layout)
    #查询会员剩余项目次数
    def load_member_projects(self):
        # 获取会员已经拥有的美容项目及其剩余次数
        query = """
            SELECT bp.name, mbp.sessions
            FROM member_beauty_projects AS mbp
            JOIN beauty_projects AS bp ON mbp.beauty_project_id = bp.id
            WHERE mbp.member_id = %s
        """
        params = (self.member_id, )
        projects = self.db_manager.execute_query(query, params)

        self.projects_table.setRowCount(0)
        for row, project in enumerate(projects):
            self.projects_table.insertRow(row)
            for col, data in enumerate(project):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable)  #高亮修改这里
                self.projects_table.setItem(row, col, item)
        # 连接cellClicked信号到highlight_project_row槽函数
        self.projects_table.cellClicked.connect(self.highlight_project_row)
        # 连接cellClicked信号到on_project_selected槽函数
        self.projects_table.cellClicked.connect(self.on_project_selected)

    def highlight_project_row(self, row):
    # 先重置所有单元格为白色
        for i in range(self.projects_table.rowCount()):
            for j in range(self.projects_table.columnCount()):
                item = self.projects_table.item(i, j)
                if item:  # 检查项目是否存在
                    item.setBackground(Qt.white)

        # 设置高亮颜色
        for j in range(self.projects_table.columnCount()):
            item = self.projects_table.item(row, j)
            if item:  # 检查项目是否存在
                item.setBackground(Qt.green)  # 更改为绿色进行测试

        self.projects_table.viewport().update()  # 强制刷新表格视图
    
    def on_project_selected(self, row):
        # 获取选中的美容项目名称
        project_name_item = self.projects_table.item(row, 0)  # 第0列是项目名称
        if project_name_item:
            project_name = project_name_item.text()

            # 将项目名称设置到项目下拉列表中
            index = self.project_combo.findText(project_name)
            if index != -1:
                self.project_combo.setCurrentIndex(index)

            # 调用update_amount()函数来更新其他相关字段
            self.update_amount()
    def get_member_info_and_discount(self):
        # 获取会员信息
        query = f"SELECT membership_level, birthday FROM members WHERE id = {self.member_id}"
        member_info = self.db_manager.execute_query(query)[0]
        member_level, member_birthday = member_info
        # 检查今天是否为会员日或会员的生日
        discount_info = self.calculate_discount(member_level, member_birthday)
        return discount_info,member_level

    def calculate_discount(self, member_level, member_birthday):
        current_day = datetime.now().day
        current_month = datetime.now().month
        is_member_birthday = (member_birthday.day == current_day) and (member_birthday.month == current_month)
        
        # 获取会员日信息和折扣
        member_day_info = self.db_manager.execute_query("SELECT day_of_month FROM member_birthday_rules WHERE creator_id = %s", (self.creator_id,))
        member_day_discount_info = self.db_manager.execute_query("SELECT member_day_discount, birthday_discount FROM membership_discounts WHERE membership_level = %s AND creator_id = %s", (member_level, self.creator_id))
        if member_day_info:
            is_member_day = member_day_info[0][0] == current_day
        else:
            is_member_day = False
        
        if member_day_discount_info:
            member_day_discount, birthday_discount = member_day_discount_info[0]
        else:
            member_day_discount = birthday_discount = 10  # 默认值
        
        discount = 1  # 默认没有折扣
        if is_member_day:
            discount = member_day_discount / 10
        elif is_member_birthday:
            discount = birthday_discount / 10
        else:
            # 如果不是会员日或生日，则查询会员等级的折扣
            discount_level = self.db_manager.execute_query("SELECT discount FROM membership_discounts WHERE membership_level = %s AND creator_id = %s", (member_level, self.creator_id))
            if discount_level:
                discount = discount_level[0][0] / 10
        
        return discount


    #读取营业员信息
    def load_clerks(self):
        query = f"SELECT name FROM clerks WHERE creator_id = {self.creator_id}"
        clerks = self.db_manager.execute_query(query)
        for clerk in clerks:
            self.clerk_combo.addItem(clerk[0])
    #读取会员美容项目
    def load_projects(self):
        query = f"SELECT id, name, single_price FROM beauty_projects WHERE creator_id = {self.creator_id}"
        projects = self.db_manager.execute_query(query)
        self.project_prices = {"无": 0}  # 初始化一个默认项目
        self.project_combo.addItem("无")  # 在列表中添加这个默认项目
        index_to_select = 0
        
        for i, project in enumerate(projects):
            self.project_combo.addItem(project[1])  # 将项目名称添加到下拉菜单中
            self.project_prices[project[1]] = project[2]  # 保存项目的价格
            if self.beauty_project_id is not None and project[0] == int(self.beauty_project_id):
                index_to_select = i + 1  # +1 因为我们在之前添加了一个"无"
        self.project_combo.setCurrentIndex(index_to_select)
    #用户选择新的项目时被调用。它首先检查会员是否已购买该项目，并且剩余次数大于0。如果是，它将消费金额设为0并禁止编辑。否则，它将消费金额设为项目的单次价格。
    def update_amount(self):
        project_name = self.project_combo.currentText()
        # 使用提取的方法获取会员信息和折扣
        discount = self.get_member_info_and_discount()[0]
        try:
            session = int(self.project_session_entry.text())
        except ValueError:
            QMessageBox.warning(self, "错误", "消耗次数必须为整数。")
            self.project_session_entry.setText('')
            return

        if project_name == "无":
            self.amount_entry.setText("0")
            self.amount_entry.setStyleSheet("color: black;")
            self.amount_entry.setReadOnly(False)
            self.actual_amount = 0
        else:
            query = f"SELECT sessions FROM member_beauty_projects WHERE member_id = {self.member_id} AND beauty_project_id = (SELECT id FROM beauty_projects WHERE name = '{project_name}' AND creator_id = '{self.creator_id}')"
            member_project = self.db_manager.execute_query(query)
            if member_project and member_project[0][0] > 0:
                self.amount_entry.setText('次卡-免费')
                self.amount_entry.setReadOnly(True)
                self.actual_amount = 0
            else:
                project_price = self.project_prices[project_name]
                if not isinstance(session, int):  # 检查session是否不是整数
                    QMessageBox.warning(self, "错误", "消耗次数必须为整数。")
                    return
                discounted_price = round(float(project_price) * discount* int(session), 2)  # 应用折扣
                self.amount_entry.setText(str(discounted_price))
                self.amount_entry.setReadOnly(False)
                self.actual_amount = discounted_price
                if discount and discount < 1:
                    self.amount_entry.setStyleSheet("color: red;")
                else:
                    self.amount_entry.setStyleSheet("color: black;")


    #确认消费
    def confirm_consume(self):
        project_name = self.project_combo.currentText()
        clerk_name = self.clerk_combo.currentText()
        member_level = self.get_member_info_and_discount()[1]
        session = int(self.project_session_entry.text())
        amount = 0
        # 获取营业员ID
        clerk_query_result = self.db_manager.execute_query("SELECT id FROM clerks WHERE name = %s", (clerk_name,))
        if not clerk_query_result:
            QMessageBox.warning(self, "错误", "请输入有效的营业员！")
            return
        clerk_id = clerk_query_result[0][0]
        self.clerk_id = clerk_id
        
        # 使用提取的方法获取会员信息和折扣
        discount = self.get_member_info_and_discount()[0]
        
        # 获取项目价格
        if project_name != "无":
            project_info = self.db_manager.execute_query("SELECT id, single_price FROM beauty_projects WHERE name = %s AND creator_id = %s", (project_name, self.creator_id))
            if not project_info:
                QMessageBox.warning(self, "错误", "项目不存在。")
                return
            project_id, project_price = project_info[0]
            
            # 检查是否使用次卡
            member_project_query = "SELECT sessions FROM member_beauty_projects WHERE member_id = %s AND beauty_project_id = %s"
            member_project = self.db_manager.execute_query(member_project_query, (self.member_id, project_id))
            if member_project and len(member_project[0]) > 0:
                old_sessions = member_project[0][0]
            else:
                old_sessions = 0
            if member_project and old_sessions > 0:
                if old_sessions- session > 0:
                    # 次卡消费逻辑
                    update_sessions_query = "UPDATE member_beauty_projects SET sessions = sessions - %s WHERE member_id = %s AND beauty_project_id = %s"
                    self.db_manager.execute_query(update_sessions_query, (session,self.member_id, project_id))
                    amount = 0  # 次卡消费，无需扣费
                elif old_sessions- session == 0:
                    delete_query = "DELETE FROM member_beauty_projects WHERE member_id = %s AND beauty_project_id = %s"
                    self.db_manager.execute_query(delete_query, (self.member_id, project_id))

                else:
                    QMessageBox.warning(self, '错误', '剩余次数不足，请检查！')
                    return
            else:
                # 正常消费逻辑
                amount = float(project_price) * discount  # 应用折扣
        else:
            amount = float(self.amount_entry.text())  # 直接从输入框获取金额
        member_info = self.db_manager.execute_query("SELECT phone,balance FROM members WHERE id = %s", (self.member_id,))[0]
        phone = member_info[0]
        project_value = f"{project_name}*{session}"
        current_balance = member_info[1]
        # 获取选择的结算方式
        payment_method = self.payment_method_combo.currentText()
        if payment_method == "账户余额":
            new_balance = current_balance - amount
            if new_balance < 0:
                QMessageBox.warning(self, "错误", "余额不足，无法完成消费。")
                return
            self.db_manager.execute_query("UPDATE members SET balance = %s WHERE id = %s", (new_balance, self.member_id))
        else:
            new_balance = None
        
        # 记录消费事务
        transaction_query = "INSERT INTO transactions (member_id, clerk_id, amount, project, transaction_date, balance ,membership_level,remark,phone, pay_method ) VALUES (%s, %s, %s, %s, NOW(), %s, %s,%s,%s,%s)"
        inserted_id  = self.db_manager.execute_insert(transaction_query, (self.member_id, clerk_id, -amount, project_value if project_name != '无' else None, new_balance, member_level,'消费',phone, payment_method ))
        id = inserted_id
        if amount == 0:
            QMessageBox.information(self, "消费成功", f"消费项目：{project_name}\n消耗次数：{session}\n剩余次数：{old_sessions-session}\n当前余额：{new_balance}")
        else:
            QMessageBox.information(self, "消费成功", f"消费项目：{project_name}\n消耗次数：{session}\n消费金额：{amount}\n当前余额：{new_balance}")
        self.accept()  # 关闭对话框或进行其他适当的UI更新
    def get_clerk_id(self):
        return self.clerk_id
    
#会员购卡
class BuyBeautyProjectDialog(QDialog):
    def __init__(self, member_id, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("购买美容项目")

        self.member_id = member_id
        self.creator_id = parent.creator_id  # 获取当前用户的ID
        # 查询当前会员的名字和余额
        query = f"SELECT name, balance FROM members WHERE id = {self.member_id}"
        member_info = self.db_manager.execute_query(query)[0]
        member_name = member_info[0]
        member_balance = member_info[1]

        # 创建标签以显示会员的名字和余额
        member_info_label = QLabel(f"会员姓名：{member_name}\n会员余额：{member_balance}元")
        member_info_label.setStyleSheet("font-weight: bold;")

        amount_label = QLabel("消费金额：")
        self.amount_entry = QLineEdit()

        sessions_label = QLabel("增加次数：")
        self.sessions_entry = QLineEdit()
        self.sessions_entry.setReadOnly(True)

        project_label = QLabel("选择美容项目：")
        self.project_combo = QComboBox()
        self.project_combo.currentIndexChanged.connect(self.update_amount)  # 刷新
        self.load_beauty_projects()

        clerk_label = QLabel("营业员：")
        self.clerk_combo = QComboBox()
        payment_method_label = QLabel("结算方式：")
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItems(["账户余额", "现付"])
        self.load_clerks()
        self.resize(650, 200)
        
        # 创建会员美容项目表格
        self.projects_table = QTableWidget()
        self.projects_table.setColumnCount(2)
        self.projects_table.setHorizontalHeaderLabels(["美容项目", "剩余次数"])
        self.projects_table.verticalHeader().setVisible(False)
        self.load_member_projects()
        
        # 创建主要布局
        layout = QVBoxLayout()
        layout.addWidget(member_info_label)
        # 创建水平布局
        hbox = QHBoxLayout()
        # 左侧布局
        left_layout = QVBoxLayout()
        left_layout.addWidget(project_label)
        left_layout.addWidget(self.project_combo)
        left_layout.addWidget(amount_label)
        left_layout.addWidget(self.amount_entry)
        left_layout.addWidget(sessions_label)
        left_layout.addWidget(self.sessions_entry)
        left_layout.addWidget(clerk_label)
        left_layout.addWidget(self.clerk_combo)
        left_layout.addWidget(payment_method_label)
        left_layout.addWidget(self.payment_method_combo)
        hbox.addLayout(left_layout)
        # 右侧布局
        right_layout = QVBoxLayout()
        purchased_label = QLabel("已购买项目")
        right_layout.addWidget(purchased_label)
        right_layout.addWidget(self.projects_table)
        hbox.addLayout(right_layout)
        
        layout.addLayout(hbox)

        confirm_button = QPushButton("确认购买")
        confirm_button.clicked.connect(self.buy_beauty_project)

        layout.addWidget(confirm_button)

        self.setLayout(layout)
    #查询会员剩余项目次数
    def load_member_projects(self):
        # 获取会员已经拥有的美容项目及其剩余次数
        query = """
            SELECT bp.name, mbp.sessions
            FROM member_beauty_projects AS mbp
            JOIN beauty_projects AS bp ON mbp.beauty_project_id = bp.id
            WHERE mbp.member_id = %s
        """
        params = (self.member_id, )
        projects = self.db_manager.execute_query(query, params)

        self.projects_table.setRowCount(0)
        for row, project in enumerate(projects):
            self.projects_table.insertRow(row)
            for col, data in enumerate(project):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsEnabled)
                self.projects_table.setItem(row, col, item)

    def update_amount(self):
        project_name = self.project_combo.currentText()
        project = self.db_manager.execute_query("SELECT card_price,times FROM beauty_projects WHERE name = %s AND creator_id = %s", (project_name, self.creator_id))[0]
        card_price = float(project[0])
        self.amount_entry.setText(str(card_price))
        self.amount_entry.setReadOnly(False)  # 允许用户修改消费金额
        sessions = int(project[1])
        self.sessions_entry.setText(str(sessions))
    #读取营业员信息
    def load_clerks(self):
        query = f"SELECT name FROM clerks WHERE creator_id = {self.creator_id}"
        clerks = self.db_manager.execute_query(query)
        for clerk in clerks:
            self.clerk_combo.addItem(clerk[0])

    #读取美容项目
    def load_beauty_projects(self):
        beauty_projects = self.db_manager.execute_query(f"SELECT name FROM beauty_projects WHERE creator_id = {self.creator_id}")
        for project in beauty_projects:
            self.project_combo.addItem(project[0])
    #购买美容项目
    def buy_beauty_project(self):
        project_name = self.project_combo.currentText()

        # 获取项目的相关信息
        project = self.db_manager.execute_query("SELECT id, times FROM beauty_projects WHERE name = %s AND creator_id = %s", (project_name, self.creator_id))[0]
        project_id = project[0]
        project_times = project[1]


        # 检查用户输入的消费金额
        try:
            card_price = float(self.amount_entry.text())
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的消费金额。")
            return
        
        # 获取当前余额
        query = f"SELECT balance,phone FROM members WHERE id = {self.member_id}"
        member = self.db_manager.execute_query(query)[0]
        current_balance = member[0]
        member_phone = member[1]
        # 获取选择的结算方式
        payment_method = self.payment_method_combo.currentText()
        if payment_method == "账户余额":
            new_balance = round(current_balance - card_price,2)
            if new_balance < 0:
                QMessageBox.warning(self, "错误", "余额不足，无法消费。")
                return
            # 更新会员余额
            query = f"UPDATE members SET balance = {new_balance}, last_consume_date = NOW() WHERE id = {self.member_id}"
            self.db_manager.execute_query(query)
        else:
            new_balance = None
        # 获取会员等级
        query = f"SELECT membership_level FROM members WHERE id = {self.member_id}"
        membership_level = self.db_manager.execute_query(query)[0][0]

        # 检查会员是否已经购买了这个项目
        member_project = self.db_manager.execute_query("SELECT * FROM member_beauty_projects WHERE member_id = %s AND beauty_project_id = %s", (self.member_id, project_id))
        if member_project:
            # 会员之前购买过这个项目，更新次数
            current_sessions = member_project[0][2]
            new_sessions = current_sessions + project_times
            self.db_manager.execute_query("UPDATE member_beauty_projects SET sessions = %s WHERE member_id = %s AND beauty_project_id = %s", (new_sessions, self.member_id, project_id))
            QMessageBox.information(self, "提示", "购买成功\n\n美容项目：{}\n次数增加：{}次\n套餐剩余：{}次\n消费金额：{}元\n剩余余额：{}元".format(project_name,project_times,new_sessions,current_balance,new_balance))
        else:
            # 会员之前没有购买过这个项目，添加新记录
            self.db_manager.execute_query("INSERT INTO member_beauty_projects (member_id, beauty_project_id, sessions, project_name, phone) VALUES (%s, %s, %s, %s, %s)", (self.member_id, project_id, project_times, project_name, member_phone))
            QMessageBox.information(self, "提示", "购买成功\n\n美容项目：{}\n套餐次数：{}次\n消费金额：{}元\n剩余余额：{}元".format(project_name,project_times,current_balance,new_balance))
        

        # 记录交易信息
        clerk_name = self.clerk_combo.currentText()
        clerk = self.db_manager.execute_query("SELECT id FROM clerks WHERE name = %s", (clerk_name,))
        if clerk is None:
            QMessageBox.warning(self, "错误", "营业员不存在或你没有权限录入此营业员的交易信息。")
            return
        clerk_id = clerk[0]


        transaction_query = f"INSERT INTO transactions (member_id, amount, project, transaction_date, balance, remark, clerk_id, membership_level, phone,pay_method) VALUES (%s,%s, %s,NOW(),%s, '购卡', %s, %s, %s, %s)"
        params =  (self.member_id, -card_price, project_name, new_balance, clerk_id, membership_level, member_phone, payment_method)
        inserted_id  = self.db_manager.execute_insert(transaction_query, params)
        self.close()

#新增美容项目对话框
class AddBeautyProjectWindow(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("新增美容项目")

        self.creator_id = parent.creator_id  # 获取当前用户的ID

        name_label = QLabel("美容项目名称：")
        self.name_entry = QLineEdit()
        times_label = QLabel("次数：")
        self.times_entry = QLineEdit()
        datetime_label = QLabel("新建时间：")
        self.datetime_entry = QDateTimeEdit(datetime.now())
        self.datetime_entry.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        single_price_label = QLabel("单次价格：")
        self.single_price_entry = QLineEdit()
        card_price_label = QLabel("次卡价格：")
        self.card_price_entry = QLineEdit()
        duration_label = QLabel("单次耗时（分钟）")
        self.duration_entry = QLineEdit()

        layout = QGridLayout()
        layout.addWidget(name_label, 0, 0)
        layout.addWidget(self.name_entry, 0, 1)
        layout.addWidget(single_price_label, 1, 0)
        layout.addWidget(self.single_price_entry, 1, 1)
        layout.addWidget(card_price_label, 2, 0)
        layout.addWidget(self.card_price_entry, 2, 1)
        layout.addWidget(times_label, 3, 0)
        layout.addWidget(self.times_entry, 3, 1)
        layout.addWidget(duration_label, 4, 0)
        layout.addWidget(self.duration_entry, 4, 1)
        layout.addWidget(datetime_label, 5, 0)
        layout.addWidget(self.datetime_entry, 5, 1)
        


        add_button = QPushButton("确定")
        add_button.clicked.connect(self.add_beauty_project)

        layout.addWidget(add_button, 6, 0, 1, 2)

        self.setLayout(layout)

    def add_beauty_project(self):
        name = self.name_entry.text()
        times = self.times_entry.text()
        single_price = self.single_price_entry.text()
        card_price = self.card_price_entry.text()
        datetime = self.datetime_entry.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        duration = self.duration_entry.text()
        if self.check_beauty_project_exists(name):
            QMessageBox.warning(self, "错误", "美容项目已存在，请输入其他美容项目。")
            return
        query = f"INSERT INTO beauty_projects (name, times, single_price, card_price, datetime, creator_id, duration) VALUES (%s, %s, %s, %s, %s, %s, %s)"
        params = (name, times, single_price, card_price, datetime, self.creator_id, duration)
        self.db_manager.execute_query(query,params)
        QMessageBox.information(self, "提示", "美容项目添加成功。")
        self.parent().search_beauty_project()
        self.close()

    def check_beauty_project_exists(self,name):
        query = f"SELECT * FROM beauty_projects WHERE name = '{name}' AND creator_id = {self.creator_id}"
        beauty_project = self.db_manager.execute_query(query)
        if beauty_project:
            return True
        return False
#修改美容项目对话框
class EditBeautyProjectWindow(QDialog):
    def __init__(self, beauty_project_id, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("修改美容项目")

        self.beauty_project_id = beauty_project_id
        self.creator_id = parent.creator_id  # 获取当前用户的ID

        query = f"SELECT * FROM beauty_projects WHERE id = {self.beauty_project_id}"
        beauty_project = self.db_manager.execute_query(query)[0]

        name_label = QLabel("美容项目名称：")
        self.name_entry = QLineEdit()
        self.name_entry.setText(beauty_project[1])
        times_label = QLabel("次数：")
        self.times_entry = QLineEdit()
        self.times_entry.setText(str(beauty_project[2]))
        single_price_label = QLabel("单次价格(元)：")
        self.single_price_entry = QLineEdit()
        self.single_price_entry.setText(str(beauty_project[6]))  # Assuming that the 4th column is single_price
        card_price_label = QLabel("次卡价格(元)")
        self.card_price_entry = QLineEdit()
        self.card_price_entry.setText(str(beauty_project[7]))
        duration_label = QLabel("单次耗时（分钟）")
        self.duration_entry = QLineEdit()
        self.duration_entry.setText(str(beauty_project[8]))

        layout = QGridLayout()
        layout.addWidget(name_label, 0, 0)
        layout.addWidget(self.name_entry, 0, 1)
        layout.addWidget(times_label, 1, 0)
        layout.addWidget(self.times_entry, 1, 1)
        layout.addWidget(single_price_label, 2, 0)
        layout.addWidget(self.single_price_entry, 2, 1)
        layout.addWidget(card_price_label, 3, 0)
        layout.addWidget(self.card_price_entry, 3, 1)
        layout.addWidget(duration_label, 4, 0)
        layout.addWidget(self.duration_entry, 4, 1)

        save_button = QPushButton("保存")
        save_button.clicked.connect(self.save_beauty_project)

        layout.addWidget(save_button, 5, 0, 1, 2)

        self.setLayout(layout)

    def save_beauty_project(self):
        name = self.name_entry.text()
        times = self.times_entry.text()
        single_price = self.single_price_entry.text()
        card_price = self.card_price_entry.text()
        duration = self.duration_entry.text()
        query = f"UPDATE beauty_projects SET name = '{name}', times = {times}, single_price = {single_price}, card_price = {card_price},duration = {duration} WHERE id = {self.beauty_project_id} AND creator_id = {self.creator_id}"
        self.db_manager.execute_query(query)
        QMessageBox.information(self, "提示", "美容项目信息修改成功。")
        self.parent().search_beauty_project()
        self.close()
#新增营业员对话框
class AddClerkWindow(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("新增营业员")

        self.creator_id = parent.creator_id  # 获取当前用户的ID

        name_label = QLabel("营业员名称：")
        self.name_entry = QLineEdit()
        phone_label = QLabel("手机号：")
        self.phone_entry = QLineEdit()

        layout = QGridLayout()
        layout.addWidget(name_label, 0, 0)
        layout.addWidget(self.name_entry, 0, 1)
        layout.addWidget(phone_label, 1, 0)
        layout.addWidget(self.phone_entry, 1, 1)

        add_button = QPushButton("确定")
        add_button.clicked.connect(self.add_clerk)

        layout.addWidget(add_button, 3, 0, 1, 2)

        self.setLayout(layout)

    def add_clerk(self):
        name = self.name_entry.text()
        phone = self.phone_entry.text()
        if self.check_clerk_exists(name):
            QMessageBox.warning(self, "错误", "营业员已存在，请输入其他营业员。")
            return
        query = f"INSERT INTO clerks (name, phone, creator_id) VALUES ('{name}','{phone}', {self.creator_id})"
        self.db_manager.execute_query(query)
        QMessageBox.information(self, "提示", "营业员添加成功。")
        self.parent().search_clerk()
        self.close()

    def check_clerk_exists(self, name):
        query = f"SELECT * FROM clerks WHERE name = '{name}' AND creator_id = {self.creator_id}"
        clerk = self.db_manager.execute_query(query)
        if clerk:
            return True
        return False
#修改营业员对话框
class ChangeClerkWindow(QDialog):
    def __init__(self, clerk_id, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.clerk_id = clerk_id
        self.setWindowTitle("修改营业员")
        query = f"SELECT name,phone FROM clerks WHERE id = {self.clerk_id}"
        clerks_info = self.db_manager.execute_query(query)[0]

        self.creator_id = parent.creator_id  # 获取当前用户的ID

        name_label = QLabel("营业员名称：")
        self.name_entry = QLineEdit()
        self.name_entry.setText(clerks_info[0])
        phone_label = QLabel("手机号：")
        self.phone_entry = QLineEdit()
        self.phone_entry.setText(clerks_info[1])

        layout = QGridLayout()
        layout.addWidget(name_label, 0, 0)
        layout.addWidget(self.name_entry, 0, 1)
        layout.addWidget(phone_label, 1, 0)
        layout.addWidget(self.phone_entry, 1, 1)

        change_button = QPushButton("保存修改")
        change_button.clicked.connect(self.change_clerk)

        layout.addWidget(change_button, 3, 0, 1, 2)

        self.setLayout(layout)

    def change_clerk(self):
        name = self.name_entry.text()
        phone = self.phone_entry.text()
        id = self.clerk_id
        query = f"UPDATE clerks SET name = '{name}', phone = '{phone}', creator_id = {self.creator_id} WHERE id = {id}"
        self.db_manager.execute_query(query)
        QMessageBox.information(self, "提示", "营业员修改成功。")
        self.parent().search_clerk()
        self.close()

    def check_clerk_exists(self, name):
        query = f"SELECT * FROM clerks WHERE name = '{name}' AND creator_id = {self.creator_id}"
        clerk = self.db_manager.execute_query(query)
        if clerk:
            return True
        return False
#设置美容床位
class BeautyBedDialog(QDialog):
    def __init__(self, parent):
        super(BeautyBedDialog, self).__init__(parent)
        self.db_manager = DatabaseManager()
        self.creator_id = parent.creator_id  # 获取当前用户的ID
        
        self.setWindowTitle("美容床位设置")
        self.resize(600, 300)  # 设置对话框尺寸
        
        # 创建布局
        self.layout = QVBoxLayout()
        
        # 创建操作按钮
        self.addButton = QPushButton("增加")
        self.deleteButton = QPushButton("删除")
        self.modifyButton = QPushButton("修改")
        self.searchButton = QPushButton("查询")
        
        # 连接按钮信号
        self.addButton.clicked.connect(self.add_bed)
        self.deleteButton.clicked.connect(self.delete_bed)
        self.modifyButton.clicked.connect(self.modify_bed)
        self.searchButton.clicked.connect(self.search_bed)
        
        # 操作按钮布局
        buttonLayout = QHBoxLayout()
        buttonLayout.addWidget(self.addButton)
        buttonLayout.addWidget(self.deleteButton)
        buttonLayout.addWidget(self.modifyButton)
        buttonLayout.addWidget(self.searchButton)
        
        # 创建表格视图
        self.table = QTableWidget()  # 创建了一个名为 self.table 的表格视图
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["id","美容床名称", "可用状态"])
        
        # 设置选择行为为整行选择
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # 设置选择模式为单个选项选择
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        
        self.table.setColumnWidth(0, 0)
        self.table.setColumnWidth(1, 200)
        self.table.setColumnWidth(2, 200)
        
        # 添加布局组件
        self.layout.addLayout(buttonLayout)
        self.layout.addWidget(self.table)  # 添加 self.table 到布局中
        
        # 设置布局
        self.setLayout(self.layout)
        
        # 加载数据
        self.search_bed()

    def add_bed(self):
        # 弹出对话框（或者使用自定义的表单窗口）让用户输入信息
        dialog = BedInfoDialog()  # 假设 BedInfoDialog 是自定义的对话框类
        if dialog.exec_() == QDialog.Accepted:
            # 获取用户输入的数据
            bed_data = dialog.get_data()
            # 构造插入数据的SQL语句
            query = "INSERT INTO beauty_beds (bed_name, status, creator_id) VALUES (%s, %s, %s)"
            params = (bed_data['bed_name'], bed_data['status'], self.creator_id)
            # 执行SQL语句
            self.db_manager.execute_query(query, params)
            # 更新表格视图
            self.search_bed()

    def delete_bed(self):
        # 获取选中的行
        selected_row = self.table.currentRow()
        
        if selected_row >= 0:
            # 获取美容床位的ID
            bed_id = self.table.item(selected_row, 0).text()
            
            # 检查是否存在引用该床位的预约记录
            check_query = "SELECT COUNT(*) FROM appointments WHERE bed_id = %s"
            params = (bed_id)
            result = self.db_manager.execute_query(check_query, params)
            
            # 如果存在引用预约记录，则提示用户不允许删除
            if result and result[0][0] > 0:
                QMessageBox.warning(self, "操作不允许", "该美容床位正在被预约使用，无法删除。")
            else:
                # 不存在引用，则执行删除操作
                delete_query = "DELETE FROM beauty_beds WHERE bed_id = %s"
                params = (bed_id)
                self.db_manager.execute_query(delete_query, params)
                # 更新表格视图
                self.search_bed()
        else:
            QMessageBox.warning(self, "警告", "请选择要删除的美容床位。")

    def modify_bed(self):
        # 获取选中的行
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            # 获取原始美容床位的信息
            bed_id = self.table.item(selected_row, 0).text()
            bed_name = self.table.item(selected_row, 1).text()
            status = self.table.item(selected_row, 2).text()
            
            # 弹出对话框（或者使用自定义的表单窗口）让用户编辑信息
            dialog = BedInfoDialog(bed_name, status)
            
            if dialog.exec_() == QDialog.Accepted:
                # 获取用户输入的更新数据
                updated_data = dialog.get_data()
                # 构造更新数据的SQL语句
                query = "UPDATE beauty_beds SET bed_name = %s, status = %s WHERE bed_id = %s"
                params = (updated_data['bed_name'], updated_data['status'], bed_id)
                # 执行SQL语句
                self.db_manager.execute_query(query, params)
                # 更新表格视图
                self.search_bed()
        else:
            QMessageBox.warning(self, "警告", "请选择要修改的美容床位。")

    # 美容床位查询功能
    def search_bed(self):
        # 构造查询SQL语句，这里假设美容床位的表格名为 beauty_beds
        # 你需要将以下SQL语句替换为与你的数据库结构相符合的查询
        query = """
            SELECT bed_id, bed_name, status
            FROM beauty_beds
            WHERE creator_id LIKE %s
            ORDER BY bed_id ASC
        """
        params = (self.creator_id)
        
        # 假设 self.db_manager 是你的数据库管理对象，你需要替换为你自己的实现
        beds = self.db_manager.execute_query(query, params)
        
        # 假设 self.search_bed_table 是你的用于显示结果的 QTableWidget
        self.table.setRowCount(0)  # 使用正确的属性名 self.table
        for row, bed in enumerate(beds):
            self.table.insertRow(row)
            for col, data in enumerate(bed):
                item = QTableWidgetItem(str(data))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # 设置单元格为可选中和启用
                self.table.setItem(row, col, item)
        
        if self.table.rowCount() > 0:
            self.table.setCurrentCell(0, 0)
class BedInfoDialog(QDialog):
    def __init__(self, bed_name='', status='', parent=None):
        super(BedInfoDialog, self).__init__(parent)
        self.setWindowTitle('输入美容床位信息')
        self.setStyleSheet("""
            QWidget {
                font: 12pt "Arial";
            }
            QLabel {
                color: #555;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px;
                border: none;
                border-radius: 5px;
                outline: none;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QTableWidget {
                gridline-color: #ccc;
            }
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 5px;
            }
            QTabWidget::pane {
                border-top: 1px solid #ccc;
            }
            QTabBar::tab {
                background-color: #f2f2f2;
                padding: 10px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
            }
            QTabBar::tab:selected {
                background-color: #ddd;
            }
        """)

        # 初始化布局
        self.layout = QVBoxLayout(self)

        # 添加床位名称输入
        self.name_layout = QHBoxLayout()
        self.name_label = QLabel('美容床名称:', self)
        self.bed_name_edit = QLineEdit(bed_name)
        self.name_layout.addWidget(self.name_label)
        self.name_layout.addWidget(self.bed_name_edit)
        self.layout.addLayout(self.name_layout)

        # 添加状态选择
        self.status_layout = QHBoxLayout()
        self.status_label = QLabel('可用状态:', self)
        self.status_combo = QComboBox(self)
        self.status_combo.addItems(['可用', '不可用'])
        if status in ['可用', '不可用']:
            self.status_combo.setCurrentText(status)
        self.status_layout.addWidget(self.status_label)
        self.status_layout.addWidget(self.status_combo)
        self.layout.addLayout(self.status_layout)

        # 创建按钮并连接信号
        self.button_layout = QHBoxLayout()
        self.ok_button = QPushButton('确定', self)
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton('取消', self)
        self.cancel_button.clicked.connect(self.reject)
        self.button_layout.addWidget(self.ok_button)
        self.button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(self.button_layout)

    def get_data(self):
        return {
            'bed_name': self.bed_name_edit.text(),
            'status': self.status_combo.currentText()
        }
#修改密码
class ChangePasswordDialog(QDialog):
    def __init__(self, parent, username):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("修改密码")

        self.username_label = QLabel("用户名：")
        self.username_entry = QLineEdit()
        self.username_entry.setReadOnly(True)
        self.username_entry.setText(username)  # 设置默认用户名
        self.old_password_label = QLabel("旧密码：")
        self.old_password_entry = QLineEdit()
        self.old_password_entry.setEchoMode(QLineEdit.Password)  # 显示为密码形式
        self.new_password_label = QLabel("新密码：")
        self.new_password_entry = QLineEdit()
        self.new_password_entry.setEchoMode(QLineEdit.Password)  # 显示为密码形式

        self.confirm_button = QPushButton("确认")
        self.confirm_button.clicked.connect(self.change_password)

        layout = QGridLayout()
        layout.addWidget(self.username_label, 0, 0)
        layout.addWidget(self.username_entry, 0, 1)
        layout.addWidget(self.old_password_label, 1, 0)
        layout.addWidget(self.old_password_entry, 1, 1)
        layout.addWidget(self.new_password_label, 2, 0)
        layout.addWidget(self.new_password_entry, 2, 1)
        layout.addWidget(self.confirm_button, 3, 0, 1, 2)

        self.setLayout(layout)


    def change_password(self):
        username = self.username_entry.text()
        old_password = self.old_password_entry.text()
        new_password = self.new_password_entry.text()


        # 从数据库中检索旧的盐值和哈希后的密码
        select_query = "SELECT password_hash, salt FROM users WHERE username = %s"
        result = self.db_manager.execute_query(select_query, (username,))[0]
        
        if result:
            password_hash, salt = result

            # 将用户输入的旧密码和数据库中的盐值进行哈希
            old_password_hash = hashlib.sha256((old_password + salt).encode('utf-8')).hexdigest()

            # 如果哈希结果和数据库中的旧密码哈希匹配，则继续下一步
            if old_password_hash == password_hash:
                # 将新密码和盐值进行哈希，并将新的哈希值保存到数据库中
                new_password_hash = hashlib.sha256((new_password + salt).encode('utf-8')).hexdigest()
                update_query = "UPDATE users SET password_hash = %s WHERE username = %s"
                self.db_manager.execute_query(update_query, (new_password_hash, username))

                QMessageBox.information(self, "提示", "密码更新成功。")
                self.close()
            else:
                QMessageBox.warning(self, "错误", "旧密码不正确。")
        else:
            QMessageBox.warning(self, "错误", "用户名不存在。")


#等级折扣设置
class DiscountSettingsDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("设置会员等级折扣")
        
        self.creator_id = self.parent().creator_id  # 从父窗口获取当前用户ID

        # UI设置
        self.levels = ["普通卡", "银卡", "金卡", "钻卡", "至尊卡"]
        self.level_entries = {}
        self.level_member_day_discount = {}
        self.level_birthday_discount = {}

        layout = QGridLayout()
        # 添加说明文字
        explanation_label = QLabel("<b><font size= 12 color='red'>提示：8.5表示85%的折扣。</font></b>")
        layout.addWidget(explanation_label, 0, 0, 1, 7)  # 这会占据前三列
        level_label = QLabel("<b><font color='red'>等级</font></b>")
        None_label = QLabel("<b><font color='red'>常规折扣</font></b>")
        member_day_discount_label = QLabel("<b><font color='red'>会员日折扣</font></b>")
        birthday_discount_label = QLabel("<b><font color='red'>生日折扣</font></b>")
        layout.addWidget(level_label, 1, 0)
        layout.addWidget(None_label, 1, 1)
        layout.addWidget(member_day_discount_label, 1, 3)
        layout.addWidget(birthday_discount_label, 1, 5)

        for index, level in enumerate(self.levels):
            label = QLabel(f"{level} ：")
            entry = QLineEdit()
            discount_label = QLabel("折")
            self.level_entries[level] = entry

            member_day_discount_entry = QLineEdit()
            member_day_discount_discount_label = QLabel("折")
            self.level_member_day_discount[level] = member_day_discount_entry

            birthday_discount_entry = QLineEdit()
            birthday_discount_discount_label = QLabel("折")
            self.level_birthday_discount[level] = birthday_discount_entry
            
            # 如果已存在的折扣值，可以在这里从数据库加载并设置为默认值
            discount = self.get_existing_discount(level)
            if discount:
                entry.setText(str(discount[0]))
                member_day_discount_entry.setText(str(discount[1]))
                birthday_discount_entry.setText(str(discount[2]))
            if level == "普通卡":
                entry.setText("无折扣")
                entry.setReadOnly(True)  # 设置为只读，不能修改
            layout.addWidget(label, index + 2, 0)  # 注意这里的index + 1，因为我们已经为说明文字占用了第0行
            layout.addWidget(entry, index + 2, 1)
            layout.addWidget(discount_label, index + 2, 2)
            layout.addWidget(member_day_discount_entry, index + 2, 3)
            layout.addWidget(member_day_discount_discount_label, index + 2, 4)
            layout.addWidget(birthday_discount_entry, index + 2, 5)
            layout.addWidget(birthday_discount_discount_label, index + 2, 6)

        self.save_button = QPushButton("保存")
        self.save_button.clicked.connect(self.save_discounts)
        layout.addWidget(self.save_button, len(self.levels) + 2, 0, 1, 7)

        self.setLayout(layout)

    def get_existing_discount(self, level):
        query = """
        SELECT discount,member_day_discount,birthday_discount FROM membership_discounts
        WHERE membership_level = %s AND creator_id = %s
        """
        result = self.db_manager.execute_query(query, (level, self.creator_id))
        if result:
            return result[0]
        else:
            return (10, 10, 10)

    def save_discounts(self):

        for level, entry in self.level_entries.items():
            discount_text = entry.text()
            member_day_discount_text = self.level_member_day_discount[level].text()
            birthday_discount_text = self.level_birthday_discount[level].text()
            
            if discount_text == '无折扣':
                discount = 10
            elif discount_text:
                discount = float(discount_text)
            else:
                discount = 10  # 默认无折扣

            if member_day_discount_text:
                member_day_discount = float(member_day_discount_text)
            else:
                member_day_discount = 10  # 默认无折扣
            
            if birthday_discount_text:
                birthday_discount = float(birthday_discount_text)
            else:
                birthday_discount = 10  # 默认无折扣
            query = """
                UPDATE membership_discounts 
                SET discount = %s, member_day_discount = %s, birthday_discount = %s
                WHERE membership_level = %s AND creator_id = %s
            """
            params = (discount, member_day_discount, birthday_discount, level,self.creator_id)
            self.db_manager.execute_query(query, params)

        QMessageBox.information(self, "提示", "折扣设置已保存。")
        self.close()


#会员等级规则设置
class LevelRuleSettingsDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("设置会员等级规则")
        
        self.creator_id = self.parent().creator_id  # 从父窗口获取当前用户ID

        # UI设置
        self.levels = ["普通卡", "银卡", "金卡", "钻卡", "至尊卡"]
        self.level_entries = {}

        layout = QGridLayout()
        # 添加说明文字
        explanation_label = QLabel("<b><font color='red'>累计充值金额达到以下数值时<br>升级到对应的等级。</font></b>")
        layout.addWidget(explanation_label, 0, 0, 1, 3)  # 这会占据前三列

        for index, level in enumerate(self.levels):
            label = QLabel(f"{level} 累计充值：")
            entry = QLineEdit()
            if level == "普通卡":
                entry.setText("0")
                entry.setReadOnly(True)  # 设置为只读，不能修改
            currency_label = QLabel("元")
            self.level_entries[level] = entry
            
            # 如果已存在的规则值，可以在这里从数据库加载并设置为默认值
            rule_value = self.get_existing_rule(level)
            if rule_value:
                entry.setText(str(rule_value))
            layout.addWidget(label, index + 1, 0)  
            layout.addWidget(entry, index + 1, 1)
            layout.addWidget(currency_label, index + 1, 2)

        self.save_button = QPushButton("保存")
        self.save_button.clicked.connect(self.save_rules)
        layout.addWidget(self.save_button, len(self.levels) + 1, 0, 1, 2)

        self.setLayout(layout)

    def get_existing_rule(self, level):
        query = """
        SELECT upgrade_amount FROM membership_level_rules 
        WHERE membership_level = %s AND creator_id = %s
        """
        result = self.db_manager.execute_query(query, (level, self.creator_id))
        if result:
            return result[0][0]
        return None

    def save_rules(self):
        for level, entry in self.level_entries.items():
            rule_value = float(entry.text())
            
            query = """
            REPLACE INTO membership_level_rules (membership_level, upgrade_amount, creator_id) 
            VALUES (%s, %s, %s)
            """
            params = (level, rule_value, self.creator_id)
            self.db_manager.execute_query(query, params)

        QMessageBox.information(self, "提示", "等级规则设置已保存。")
        self.close()
#会员日设置
class BirthdayRulesDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.setWindowTitle("会员日设置")
        
        self.creator_id = self.parent().creator_id  # 从父窗口获取当前用户ID

        # UI设置
        layout = QGridLayout()
        explanation_label = QLabel("<b><font color='red'>设置每月的会员日：</font></b>")
        layout.addWidget(explanation_label, 0, 0, 1, 4)  # 这会占据前四列

        day_label = QLabel("会员日：")
        self.day_entry = QSpinBox()
        day = QLabel("日")  
        self.day_entry.setRange(1, 31) # 设置日期范围为1到31
        

        # 从数据库加载已存在的设置并作为默认值
        existing_day = self.get_existing_member_day_and_discount()
        if existing_day:
            self.day_entry.setValue(existing_day)
        
        layout.addWidget(day_label, 1, 0)
        layout.addWidget(self.day_entry, 1, 1)
        layout.addWidget(day, 1, 2)
   

        self.save_button = QPushButton("保存")
        self.save_button.clicked.connect(self.save_member_day_and_discount)
        layout.addWidget(self.save_button, 2, 0, 1, 4)

        self.setLayout(layout)

    def get_existing_member_day_and_discount(self):
        query = """
        SELECT day_of_month FROM member_birthday_rules
        WHERE creator_id = %s
        """
        result = self.db_manager.execute_query(query, (self.creator_id,))
        if result:
            return result[0][0]
        return None

    def save_member_day_and_discount(self):
        member_day = self.day_entry.value()

        
        query = """
        INSERT INTO member_birthday_rules (day_of_month,  creator_id)
        VALUES (%s,  %s)
        ON DUPLICATE KEY UPDATE day_of_month = %s
        """
        params = (member_day, self.creator_id, member_day)

        self.db_manager.execute_query(query, params)

        QMessageBox.information(self, "提示", "会员日设置已保存。")
        self.close()
#会员预约表
class AppointmentViewerTab(QWidget):
    def __init__(self, db_manager, creator_id):  # 加入 creator_id 作为参数
        super().__init__()
        self.db_manager = db_manager
        self.creator_id = creator_id  # 初始化 creator_id
        self.appointment_id = None
        self.selected_date = QDate.currentDate()
        self.phone = None
        self.initUI()
    def highlight_row(self):
        for i in range(self.tableWidget.rowCount()): 
            appointment_status_item = self.tableWidget.item(i, 7)  # 假设预约状态在第8列 (列索引从0开始)
            for j in range(self.tableWidget.columnCount()):
                if appointment_status_item.text() == "待确认":
                    item = self.tableWidget.item(i, j)
                    if item:  
                        item.setBackground(Qt.white)
                        item.setForeground(Qt.red)  # 将字体设置为红色
                else:
                    item = self.tableWidget.item(i, j)
                    if item:  
                        item.setBackground(Qt.white)

        selected_rows = self.tableWidget.selectionModel().selectedRows()
        if selected_rows:
            current_row = selected_rows[0].row()

            appointment_status_item = self.tableWidget.item(current_row, 7)  
            if appointment_status_item:
                if appointment_status_item.text() == "已确认":
                    self.confirm_appointment_btn.setEnabled(False)
                    self.confirm_appointment_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    self.cancel_appointment_btn.setEnabled(True)
                    self.cancel_appointment_btn.setStyleSheet("")  # 恢复按钮默认样式
                    self.complete_service_btn.setEnabled(True)
                    self.complete_service_btn.setStyleSheet("")  # 恢复按钮默认样式
                    for j in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(current_row, j)
                        if item:  
                            item.setBackground(Qt.gray)  
                elif appointment_status_item.text() == "已完成":
                    self.confirm_appointment_btn.setEnabled(False)
                    self.confirm_appointment_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    self.cancel_appointment_btn.setEnabled(False)
                    self.cancel_appointment_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    self.complete_service_btn.setEnabled(False)
                    self.complete_service_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    for j in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(current_row, j)
                        if item:  
                            item.setBackground(Qt.gray)
                elif appointment_status_item.text() == "已取消":
                    self.confirm_appointment_btn.setEnabled(False)
                    self.confirm_appointment_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    self.cancel_appointment_btn.setEnabled(False)
                    self.cancel_appointment_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    self.complete_service_btn.setEnabled(False)
                    self.complete_service_btn.setStyleSheet("background-color: gray")  # 将按钮置灰
                    for j in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(current_row, j)
                        if item:
                            item.setBackground(Qt.gray)
                else:
                    self.confirm_appointment_btn.setEnabled(True)
                    self.confirm_appointment_btn.setStyleSheet("")  # 恢复按钮默认样式
                    self.cancel_appointment_btn.setEnabled(True)
                    self.cancel_appointment_btn.setStyleSheet("")  # 恢复按钮默认样式
                    self.complete_service_btn.setEnabled(True)
                    self.complete_service_btn.setStyleSheet("")  # 恢复按钮默认样式
                    for j in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(current_row, j)
                        if item:
                            item.setBackground(Qt.green)

        self.tableWidget.viewport().update()

    def initUI(self):
        # 主垂直布局
        layout = QVBoxLayout()

        # 添加选择日期标签和日期选择器
        layout.addWidget(QLabel("选择日期查看预约情况："))

        # 水平布局 - 用于放置日期选择器和按钮
        button_layout = QHBoxLayout()
        self.date_picker = QDateEdit(self)
        self.date_picker.setCalendarPopup(True)
        self.date_picker.setDate(QDate.currentDate())
        button_layout.addWidget(self.date_picker)
        
        self.load_btn = QPushButton("加载预约", self)
        self.load_btn.clicked.connect(self.load_appointments)
        button_layout.addWidget(self.load_btn)

        # 修改"预约确认"按钮为类变量
        self.confirm_appointment_btn = QPushButton("预约确认", self)
        self.confirm_appointment_btn.clicked.connect(self.confirm_appointment)
        button_layout.addWidget(self.confirm_appointment_btn)
        
        self.cancel_appointment_btn = QPushButton("取消预约", self)
        self.cancel_appointment_btn.clicked.connect(self.cancel_appointment)
        button_layout.addWidget(self.cancel_appointment_btn)
            
        # 添加完成服务按钮
        self.complete_service_btn = QPushButton("完成服务", self)
        self.complete_service_btn.clicked.connect(self.complete_service)  # 这里需要定义一个新的槽函数来处理完成服务的事件
        button_layout.addWidget(self.complete_service_btn)

        # 将水平布局添加到主垂直布局中
        layout.addLayout(button_layout)

        # 数据表格
        self.tableWidget = QTableWidget(self)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 禁止编辑
        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        # 当表格中的选定项发生变化时，调用 highlight_row 方法
        self.tableWidget.itemSelectionChanged.connect(self.highlight_row)
        self.tableWidget.doubleClicked.connect(self.complete_service) #双击事件
        self.tableWidget.setColumnCount(12)  # 增加一列来显示预约状态
        self.tableWidget.setHorizontalHeaderLabels(["预约ID","会员ID","会员名", "手机号", "美容床位", "项目", "时间", "预约状态", "项目ID","特殊要求", "营业员","完成时间"])  # 添加预约状态的列标题
        layout.addWidget(self.tableWidget)
        
        self.tableWidget.setColumnWidth(0, 0) 
        self.tableWidget.setColumnWidth(1, 0) 
        self.tableWidget.setColumnWidth(8, 0) 
        self.tableWidget.setColumnWidth(9,240) 
        self.tableWidget.setColumnWidth(11,240) 
        self.setLayout(layout)

    # 预约查询
    def load_appointments(self):
        self.highlight_row() 
        self.selected_date = self.date_picker.date()
        # 将选定的日期转换为适合SQL查询的字符串格式
        date_str = self.selected_date.toString("yyyy-MM-dd")
        
        # 修改查询来获取预约状态
        query = """
        SELECT appointments.appointment_id, members.id, members.name, members.phone, beauty_beds.bed_name, 
            beauty_projects.name, appointments.start_time, appointments.end_time, appointments.status,appointments.project_id, clerks.name, finish_data,special_req
        FROM appointments
        JOIN members ON appointments.member_id = members.id
        JOIN beauty_beds ON appointments.bed_id = beauty_beds.bed_id
        JOIN beauty_projects ON appointments.project_id = beauty_projects.id
        LEFT JOIN clerks ON appointments.clerk_id = clerks.id
        WHERE DATE(appointments.start_time) = %s AND appointments.creator_id = %s
        ORDER BY appointments.start_time ASC
        """

        results = self.db_manager.execute_query(query, (date_str,self.creator_id))
        
        # 将结果加载到表部件中
        self.tableWidget.setRowCount(len(results))
        for row, result in enumerate(results):
            items = [
                result[0],  # 预约ID
                result[1],  # 会员ID
                result[2],  # 会员姓名
                result[3],  # 会员手机号
                result[4],  # 床位名称
                result[5],  # 项目名称
                f"{result[6].strftime('%H:%M')} - {result[7].strftime('%H:%M')}",  # 开始和结束时间
                result[8],  # 预约状态
                result[9],  #项目ID
                result[12],   #特殊要求
                result[10],  #营业员
                result[11]  #完成时间
                
            ]
            for col, item in enumerate(items):
                self.tableWidget.setItem(row, col, QTableWidgetItem(str(item)))
        self.highlight_row()

    #确认预约
    def confirm_appointment(self):
        # 获取选中的预约
        selected_row = self.tableWidget.currentRow()
        if selected_row != -1:
            appointment_id = self.tableWidget.item(selected_row, 0).text()
            phone = self.tableWidget.item(selected_row, 3).text()
            reply = QMessageBox.question(self, "确认预约", "确定要确认选中的预约吗？",
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return
            
            # 使用预约ID更新数据库中的状态为"已确认"
            query = "UPDATE appointments SET status = '已确认' WHERE appointment_id = %s AND creator_id = %s"
            self.db_manager.execute_query(query, (appointment_id, self.creator_id))
        
        else:
            QMessageBox.warning(self, "提示", "请先选择一个预约！")
            return
        # 重新加载预约以显示更新的状态
        self.load_appointments()
    #取消预约
    def cancel_appointment(self):
        # 获取选中的预约
        selected_row = self.tableWidget.currentRow()
        if selected_row != -1:
            appointment_id = self.tableWidget.item(selected_row, 0).text()
            phone = self.tableWidget.item(selected_row, 3).text()
            # 确认是否真的要取消预约
            reply = QMessageBox.question(self, "取消预约", "确定要取消选中的预约吗？",
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return
            # 使用预约ID更新数据库中的状态为"已取消"
            query = "UPDATE appointments SET status = '已取消' WHERE appointment_id = %s AND creator_id = %s"
            self.db_manager.execute_query(query, (appointment_id, self.creator_id))
        else:
            QMessageBox.warning(self, "提示", "请先选择一个预约！")
            return
        # 重新加载预约以显示更新的状态
        self.load_appointments()
    # 新增：完成服务槽函数
    def complete_service(self):
        # 获取选中的预约
        selected_row = self.tableWidget.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "提示", "请选择要完成的预约!")
            return
        else:
            appointment_id = self.tableWidget.item(selected_row, 0).text()
            member_id = self.tableWidget.item(selected_row, 1).text() # 第1列是会员ID
            phone = self.tableWidget.item(selected_row, 3).text()
            beauty_project_id = self.tableWidget.item(selected_row, 8).text()
            status = self.tableWidget.item(selected_row, 7).text()
            if "已完成" in status:
                QMessageBox.warning(self, "警告", "服务已完成!")
                return
            else:
            # 弹出消费对话框
                consume_dialog = ConsumeDialog(member_id, appointment_id, self, beauty_project_id, from_complete_service=True)  # 传入会员ID作为参数
                result = consume_dialog.exec_()

        if result == QDialog.Accepted:  # 如果用户在消费对话框中点击了确定或保存按钮
            clerk_id = consume_dialog.get_clerk_id()
            # 更新数据库中的预约状态
            query = "UPDATE appointments SET status='已完成',clerk_id = %s ,finish_data = NOW() WHERE appointment_id=%s"
            self.db_manager.execute_query(query, (clerk_id, appointment_id,))

            # 重新加载预约，以反映更改
            self.load_appointments()
            QMessageBox.information(self, "提示", "服务已完成!")
            

#营业时间设置
class BusinessHoursDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.creator_id = self.parent().creator_id  # 从父窗口获取当前用户ID
        self.setWindowTitle("营业时间设置")
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()
        grid_layout = QGridLayout()

        # 为每天创建时间选择器
        days = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        self.time_editors = {}
        
        for index, day in enumerate(days):
            label = QLabel(day)
            start_time_edit = QTimeEdit()
            end_time_edit = QTimeEdit()

            # 设置默认值或从数据库获取值
            start_time, end_time = self.get_business_hours_from_db(day)
            
            # 检查从数据库返回的时间是否为None
            if start_time:
                start_time = (datetime.min + start_time).time()
                start_time_edit.setTime(start_time)
                
            if end_time:
                end_time = (datetime.min + end_time).time()
                end_time_edit.setTime(end_time)

            grid_layout.addWidget(label, index, 0)
            grid_layout.addWidget(start_time_edit, index, 1)
            grid_layout.addWidget(end_time_edit, index, 2)

            self.time_editors[day] = (start_time_edit, end_time_edit)

        layout.addLayout(grid_layout)

        # 添加保存按钮
        save_button = QPushButton("保存")
        save_button.clicked.connect(self.save_business_hours)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def get_business_hours_from_db(self, day):
        # 从数据库获取营业时间
        # 返回开始和结束时间
        query = """
        SELECT start_time, end_time
        FROM business_hours
        WHERE day = %s AND creator_id = %s
        """
        results = self.db_manager.execute_query(query, (day, self.creator_id))
        if results:
            return results[0][0], results[0][1]
        return None, None  # 默认值

    def save_business_hours(self):
        # 将设置保存到数据库
        for day, (start_time_edit, end_time_edit) in self.time_editors.items():
            start_time = start_time_edit.time()
            end_time = end_time_edit.time()

            # 先获取要更新记录的 id
            get_id_query = """
            SELECT id
            FROM business_hours
            WHERE day = %s AND creator_id = %s
            """
            result = self.db_manager.execute_query(get_id_query, (day, self.creator_id))
            if result:
                # 如果记录存在，更新它
                business_hour_id = result[0][0]
                update_query = """
                UPDATE business_hours
                SET start_time = %s, end_time = %s
                WHERE id = %s
                """
                self.db_manager.execute_query(update_query, (start_time.toString(), end_time.toString(), business_hour_id))
            else:
                # 如果记录不存在，插入新的记录
                insert_query = """
                INSERT INTO business_hours(day, start_time, end_time, creator_id)
                VALUES(%s, %s, %s, %s)
                """
                self.db_manager.execute_query(insert_query, (day, start_time.toString(), end_time.toString(), self.creator_id))

        self.accept()
        QMessageBox.information(self, "提示", "保存成功！")
        self.close()
    
#链接数据库
class DatabaseManager:
    def __init__(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        
        self.host = config['database']['host']
        self.user = config['database']['user']
        self.password = config['database']['password']
        self.db = config['database']['db']
        self.connect()  # 在初始化时调用连接方法

    def connect(self):
        """建立到数据库的新连接。"""
        self.conn = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.db, connect_timeout=60)
        self.conn.autocommit(True)

    def ensure_connection(self):
        """确保连接处于活动状态，如果需要则重新连接。"""
        try:
            self.conn.ping(reconnect=True)
        except:
            self.connect()  # 如果ping失败，则重新连接

    def __del__(self):
        self.conn.close()

    def begin_transaction(self):
        self.conn.autocommit(False)

    def commit_transaction(self):
        self.conn.commit()
        self.conn.autocommit(True)

    def rollback_transaction(self):
        self.conn.rollback()
        self.conn.autocommit(True)

    def execute_query(self, query, params=()):
        self.ensure_connection()  # 在执行查询之前确保连接处于活动状态
        cursor = self.conn.cursor()
        try:
            cursor.execute(query, params)
            result = cursor.fetchall()
            return result if result is not None else []
        except pymysql.err.OperationalError as e:
            # 处理错误，可能通过重新连接并重试查询
            self.connect()
            # 您可以选择在此处重试查询，或只是引发错误以通知调用者
            raise e
        finally:
            cursor.close()
            if not self.conn.get_autocommit():
                self.conn.commit()  # 当在事务中时，确保每个查询的更改都被提交
    #用于使用lastrowid获取新插入记录的ID
    def execute_insert(self, query, params=()):
        self.ensure_connection()  # 在执行查询之前确保连接处于活动状态
        cursor = self.conn.cursor()
        try:
            cursor.execute(query, params)
            self.conn.commit()  # 确保更改被提交
            return cursor.lastrowid  # 返回新插入记录的ID
        except pymysql.err.OperationalError as e:
            # 处理错误，可能通过重新连接并重试查询
            self.connect()
            raise e
        finally:
            cursor.close()

#加载QSS文件
def load_stylesheet(self):
    try:
        with open("style.qss", "r", encoding='utf-8') as file:
            self.setStyleSheet(file.read())
    except Exception as e:
        print(f"Failed to load stylesheet: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    latest_version = check_for_updates()
    if latest_version:
        response = QMessageBox.question(None, 'Update Available', f'最新版本 {latest_version} 的应用程序可供下载。您要下载并更新吗?',QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if response == QMessageBox.Yes:
            # 创建一个进度对话框
            progress = QProgressDialog("正在下载更新...", "禁止取消", 0, 100)
            progress.setAutoClose(True)
            progress.setModal(True)
            progress.setWindowTitle("更新进度")

            def update_progress(ratio):
                progress.setValue(int(ratio * 100))
                if progress.wasCanceled():
                    return False
                return True

            if download_new_version(latest_version, update_progress):
                progress.close()
                QMessageBox.information(None, 'Update Downloaded',
                                        '新版本已下载完成。请重新启动应用程序以应用更新。',
                                        QMessageBox.Ok)
            else:
                progress.close()
                QMessageBox.critical(None, 'Update Failed', '下载新版本失败。请稍后再试。',QMessageBox.Ok)
    else:
        load_stylesheet(app)
        login_window = LoginWindow()
        login_window.show()
        sys.exit(app.exec_())